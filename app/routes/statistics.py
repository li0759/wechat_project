from flask import Blueprint, jsonify, request, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity
from ..models import User, Club, ClubMember, Event, EventJoin, ClubFee, PayGroup, Moment
from .. import db
from app.permission import check_permission, statistics
from datetime import datetime, timedelta
import io
import tempfile
import os
import re
import shutil
import subprocess
import zipfile
from collections import defaultdict
from sqlalchemy import case
from minio import Minio
from minio.error import S3Error
from flask import current_app
import requests
import hashlib
import mimetypes
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as ExcelImage
    # 尝试导入Pillow，这是openpyxl图片功能所需的
    try:
        from PIL import Image as PILImage
        PILLOW_AVAILABLE = True
    except ImportError:
        PILLOW_AVAILABLE = False
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    PILLOW_AVAILABLE = False

bp = Blueprint('statistics', __name__, url_prefix='/api/v1/statistics')

def get_minio_client():
    """获取MinIO客户端"""
    try:
        return Minio(
            current_app.config['MINIO_ENDPOINT'],
            access_key=current_app.config['MINIO_ACCESS_KEY'],
            secret_key=current_app.config['MINIO_SECRET_KEY'],
            secure=current_app.config['MINIO_SECURE']
        )
    except Exception as e:
        current_app.logger.error(f"MinIO客户端初始化失败: {str(e)}")
        raise

def ensure_bucket_exists(minio_client, bucket_name):
    """确保bucket存在"""
    try:
        if not minio_client.bucket_exists(bucket_name):
            minio_client.make_bucket(bucket_name)
    except Exception as e:
        current_app.logger.error(f"创建bucket失败: {str(e)}")
        raise

@bp.route('/export/all_club/users', methods=['GET'])
@jwt_required()
def export_all_club_users():
    """导出所有协会会员详情（每协会一个Excel+文件夹，整体压缩包）"""
    # 权限检查
    has_permission, message = check_permission(statistics.export_all_club_users.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200
    
    split_by_club_raw = request.args.get('split_by_club', '1')
    split_by_club = str(split_by_club_raw).lower() in {'1', 'true', 'yes', 'on'}
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    activity_start_date = request.args.get('activity_start_date', '').strip()
    activity_end_date = request.args.get('activity_end_date', '').strip()

    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        activity_start_dt, activity_end_dt = parse_date_range(activity_start_date, activity_end_date)
        clubs = Club.query.filter_by(isDelete=False).all()
        if split_by_club:
            return create_all_club_users_archive(clubs, 'all_club_users', start_dt, end_dt, activity_start_dt, activity_end_dt)
        return create_all_users_single_archive(clubs, 'all_club_users', start_dt, end_dt, activity_start_dt, activity_end_dt)
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'导出失败: {str(e)}'}), 200

@bp.route('/export/all_club/users/wecom_media', methods=['GET'])
@jwt_required()
def export_all_club_users_wecom_media():
    """导出所有协会用户，并上传为企业微信会话文件素材，返回 media_id。"""
    has_permission, message = check_permission(statistics.export_all_club_users.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200

    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200

    split_by_club_raw = request.args.get('split_by_club', '1')
    split_by_club = str(split_by_club_raw).lower() in {'1', 'true', 'yes', 'on'}
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    activity_start_date = request.args.get('activity_start_date', '').strip()
    activity_end_date = request.args.get('activity_end_date', '').strip()

    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        activity_start_dt, activity_end_dt = parse_date_range(activity_start_date, activity_end_date)
        clubs = Club.query.filter_by(isDelete=False).all()
        export_response = (
            create_all_club_users_archive(clubs, 'all_club_users', start_dt, end_dt, activity_start_dt, activity_end_dt)
            if split_by_club else
            create_all_users_single_archive(clubs, 'all_club_users', start_dt, end_dt, activity_start_dt, activity_end_dt)
        )
        export_payload = export_response.get_json(silent=True) or {}

        if export_payload.get('code') != 200:
            return jsonify({
                'code': export_payload.get('code', 5000),
                'message': export_payload.get('message', '导出失败')
            }), 200

        file_data = export_payload.get('data') or {}
        object_path = file_data.get('file_path')
        file_name = file_data.get('filename') or f"all_club_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if not object_path:
            return jsonify({'code': 5000, 'message': '导出结果缺少文件路径'}), 200

        media_id = upload_minio_object_to_wecom_media(object_path=object_path, file_name=file_name)
        return jsonify({
            'code': 200,
            'message': '已生成企业微信会话文件',
            'data': {
                'media_id': media_id,
                'filename': file_name,
                'archive_format': file_data.get('archive_format'),
                'split_by_club': split_by_club,
                'start_date': start_dt.strftime('%Y-%m-%d') if start_dt else '',
                'end_date': end_dt.strftime('%Y-%m-%d') if end_dt else '',
                'activity_start_date': activity_start_dt.strftime('%Y-%m-%d') if activity_start_dt else '',
                'activity_end_date': activity_end_dt.strftime('%Y-%m-%d') if activity_end_dt else ''
            }
        }), 200
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'生成企业微信会话文件失败: {str(e)}'}), 200

@bp.route('/wecom/send_media_to_self', methods=['POST'])
@jwt_required()
def send_wecom_media_to_self():
    """无会话上下文时，将 media_id 文件发给当前登录用户本人。"""
    has_permission, message = check_permission(statistics.export_all_club_users.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    payload = request.get_json(silent=True) or {}
    media_id = (payload.get('media_id') or '').strip()
    if not media_id:
        return jsonify({'code': 4001, 'message': '缺少 media_id'}), 200

    user_id = get_jwt_identity()
    user = User.query.filter_by(userID=user_id).first()
    if not user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    if not user.wecomUserID:
        return jsonify({'code': 4005, 'message': '当前用户未绑定企业微信账号，无法发送'}), 200

    try:
        send_wecom_file_message_to_users([user.wecomUserID], media_id)
        return jsonify({'code': 200, 'message': '已发送到你本人的企业微信'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'发送失败: {str(e)}'}), 200

@bp.route('/export/club/<int:club_id>/all_event/details/wecom_media', methods=['GET'])
@jwt_required()
def export_club_all_event_details_wecom_media(club_id):
    """导出单协会活动详情并上传企业微信素材，返回 media_id。"""
    has_permission, message = check_permission(statistics.export_club_all_event_details.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200

    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200
    if not PILLOW_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装图片处理库'}), 200

    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    try:
        club = Club.query.filter_by(clubID=club_id).first()
        if not club:
            return jsonify({'code': 4004, 'message': '协会不存在'}), 200

        start_dt, end_dt = parse_date_range(start_date, end_date)
        club_events = Event.query.filter_by(clubID=club_id).all()
        events = [event for event in club_events if is_event_in_range(event, start_dt, end_dt)]

        export_response = create_all_event_details_single_package(
            events=events,
            filename_prefix=f'club_{club_id}_all_events',
            start_date=start_dt,
            end_date=end_dt
        )
        export_payload = export_response.get_json(silent=True) or {}
        if export_payload.get('code') != 200:
            return jsonify({
                'code': export_payload.get('code', 5000),
                'message': export_payload.get('message', '导出失败')
            }), 200

        file_data = export_payload.get('data') or {}
        object_path = file_data.get('file_path')
        file_name = file_data.get('filename') or f"club_{club_id}_all_events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if not object_path:
            return jsonify({'code': 5000, 'message': '导出结果缺少文件路径'}), 200

        media_id = upload_minio_object_to_wecom_media(object_path=object_path, file_name=file_name)
        return jsonify({
            'code': 200,
            'message': '已生成企业微信文件素材',
            'data': {
                'media_id': media_id,
                'filename': file_name,
                'archive_format': file_data.get('archive_format'),
                'start_date': start_dt.strftime('%Y-%m-%d') if start_dt else '',
                'end_date': end_dt.strftime('%Y-%m-%d') if end_dt else ''
            }
        }), 200
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'生成企业微信会话文件失败: {str(e)}'}), 200

@bp.route('/export/all_club/all_event/details/wecom_media', methods=['GET'])
@jwt_required()
def export_all_club_all_event_details_wecom_media():
    """导出所有协会活动详情并上传企业微信素材，返回 media_id。"""
    has_permission, message = check_permission(statistics.export_all_club_all_event_details.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200

    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200
    if not PILLOW_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装图片处理库'}), 200

    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    split_by_club_raw = request.args.get('split_by_club', '1')
    split_by_club = str(split_by_club_raw).lower() in {'1', 'true', 'yes', 'on'}

    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        all_events = Event.query.join(Club).all()
        events = [event for event in all_events if is_event_in_range(event, start_dt, end_dt)]

        export_response = (
            create_all_club_event_details_rar_package(
                events=events,
                filename_prefix='all_club_all_events',
                start_date=start_dt,
                end_date=end_dt
            ) if split_by_club else create_all_event_details_single_package(
                events=events,
                filename_prefix='all_club_all_events',
                start_date=start_dt,
                end_date=end_dt
            )
        )
        export_payload = export_response.get_json(silent=True) or {}
        if export_payload.get('code') != 200:
            return jsonify({
                'code': export_payload.get('code', 5000),
                'message': export_payload.get('message', '导出失败')
            }), 200

        file_data = export_payload.get('data') or {}
        object_path = file_data.get('file_path')
        file_name = file_data.get('filename') or f"all_club_all_events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if not object_path:
            return jsonify({'code': 5000, 'message': '导出结果缺少文件路径'}), 200

        media_id = upload_minio_object_to_wecom_media(object_path=object_path, file_name=file_name)
        return jsonify({
            'code': 200,
            'message': '已生成企业微信文件素材',
            'data': {
                'media_id': media_id,
                'filename': file_name,
                'archive_format': file_data.get('archive_format'),
                'split_by_club': split_by_club
            }
        }), 200
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'生成企业微信会话文件失败: {str(e)}'}), 200


@bp.route('/export/club/<int:club_id>/member_activity/wecom_media', methods=['GET'])
@jwt_required()
def export_club_member_activity_wecom_media(club_id):
    """导出成员参与明细并上传企业微信素材。user_ids 支持单个或逗号分隔多个，如 user_ids=5 或 user_ids=5,8。"""
    has_permission, message = check_permission(statistics.export_all_club_users.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200

    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200

    club = Club.query.filter_by(clubID=club_id).first()
    if not club:
        return jsonify({'code': 4004, 'message': '协会不存在'}), 200

    raw_user_ids = (request.args.get('user_ids') or '').strip()
    if not raw_user_ids:
        return jsonify({'code': 4001, 'message': '缺少 user_ids 参数'}), 200

    picked_ids = []
    for token in raw_user_ids.split(','):
        sid = token.strip()
        if not sid:
            continue
        if sid.isdigit():
            picked_ids.append(int(sid))
    picked_ids = list(dict.fromkeys(picked_ids))
    if not picked_ids:
        return jsonify({'code': 4001, 'message': 'user_ids 参数无效'}), 200

    members = (
        ClubMember.query.filter(
            ClubMember.clubID == club_id,
            ClubMember.userID.in_(picked_ids),
            ClubMember.isDelete == False,
        )
        .all()
    )
    if not members:
        return jsonify({'code': 4004, 'message': '未找到可导出的成员'}), 200

    member_by_uid = {m.userID: m for m in members}
    sorted_user_ids = [uid for uid in picked_ids if uid in member_by_uid]

    event_joins = (
        EventJoin.query.join(Event, Event.eventID == EventJoin.eventID)
        .filter(
            Event.clubID == club_id,
            EventJoin.userID.in_(sorted_user_ids),
            EventJoin.isDelete == False,
        )
        .order_by(EventJoin.joinDate.desc(), EventJoin.joinID.desc())
        .all()
    )

    join_map = defaultdict(list)
    event_ids = set()
    for ej in event_joins:
        join_map[ej.userID].append(ej)
        if ej.eventID is not None:
            event_ids.add(ej.eventID)

    moment_map = defaultdict(list)
    if event_ids:
        moments = (
            Moment.query.filter(
                Moment.creatorID.in_(sorted_user_ids),
                Moment.ref_event_ID.in_(list(event_ids)),
            )
            .order_by(Moment.createDate.desc(), Moment.momentID.desc())
            .all()
        )
        for m in moments:
            if m.ref_event_ID is None:
                continue
            key = (m.creatorID, m.ref_event_ID)
            moment_map[key].append(m)

    def _fmt(dt):
        if not dt:
            return ''
        return dt.strftime('%Y-%m-%d %H:%M')

    headers = ['用户姓名', '活动名称', '参加时间', '打卡时间', '人员发布的动态']
    data_rows = []
    moment_files_per_row = []
    avatar_file_per_row = []
    cover_file_per_row = []
    for uid in sorted_user_ids:
        member = member_by_uid[uid]
        user_name = member.user.userName if member.user else str(uid)
        user_avatar = member.user.avatar if member.user else None
        joins = join_map.get(uid, [])
        if not joins:
            data_rows.append([user_name, '无活动记录', '', '', ''])
            moment_files_per_row.append([])
            avatar_file_per_row.append(user_avatar)
            cover_file_per_row.append(None)
            continue
        for ej in joins:
            event_title = ej.event.title if ej.event else '未知活动'
            event_cover = ej.event.cover if ej.event else None
            user_moments = moment_map.get((uid, ej.eventID), [])
            moment_lines = []
            for m in user_moments:
                desc = (m.description or '').strip() or '（无文字）'
                t = _fmt(m.createDate)
                moment_lines.append(f'{desc}（{t}）' if t else desc)
            data_rows.append([
                user_name,
                event_title,
                _fmt(ej.joinDate),
                _fmt(ej.clockinDate) or '未打卡',
                '；'.join(moment_lines) if moment_lines else '无',
            ])
            moment_files_per_row.append(collect_moment_files_from_moments(user_moments))
            avatar_file_per_row.append(user_avatar)
            cover_file_per_row.append(event_cover)

    if not PILLOW_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装图片处理库，无法打包动态原图'}), 200

    try:
        export_resp = create_participation_export_archive(
            headers=headers,
            data_rows=data_rows,
            moment_files_per_row=moment_files_per_row,
            filename_prefix=f'club_{club_id}_member_activity',
            excel_basename='成员活动明细',
            sheet_title='成员活动明细',
            avatar_file_per_row=avatar_file_per_row,
            cover_file_per_row=cover_file_per_row,
            export_layout='member_activity',
        )
        export_payload = export_resp.get_json(silent=True) or {}
        if export_payload.get('code') != 200:
            return jsonify({
                'code': export_payload.get('code', 5000),
                'message': export_payload.get('message', '导出失败'),
            }), 200

        file_data = export_payload.get('data') or {}
        object_path = file_data.get('file_path')
        file_name = file_data.get('filename') or f"club_{club_id}_member_activity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if not object_path:
            return jsonify({'code': 5000, 'message': '导出结果缺少文件路径'}), 200

        media_id = upload_minio_object_to_wecom_media(object_path=object_path, file_name=file_name)
        return jsonify({
            'code': 200,
            'message': '已生成企业微信文件素材（含动态原图压缩包）',
            'data': {
                'media_id': media_id,
                'filename': file_name,
                'archive_format': file_data.get('archive_format', 'zip'),
                'club_id': club_id,
                'selected_user_count': len(sorted_user_ids),
                'moment_image_count': file_data.get('moment_image_count', 0),
            },
        }), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'生成企业微信会话文件失败: {str(e)}'}), 200

@bp.route('/export/club/<int:club_id>/event_participation/wecom_media', methods=['GET'])
@jwt_required()
def export_club_event_participation_wecom_media(club_id):
    """导出活动参与明细并上传企业微信素材。event_ids 支持单个或逗号分隔多个，如 event_ids=12 或 event_ids=12,34。"""
    has_permission, message = check_permission(statistics.export_club_event_participation.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200

    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200

    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200

    club = Club.query.filter_by(clubID=club_id).first()
    if not club:
        return jsonify({'code': 4004, 'message': '协会不存在'}), 200

    raw_event_ids = (request.args.get('event_ids') or '').strip()
    if not raw_event_ids:
        return jsonify({'code': 4001, 'message': '缺少 event_ids 参数'}), 200

    picked_ids = []
    for token in raw_event_ids.split(','):
        sid = token.strip()
        if sid.isdigit():
            picked_ids.append(int(sid))
    picked_ids = list(dict.fromkeys(picked_ids))
    if not picked_ids:
        return jsonify({'code': 4001, 'message': 'event_ids 参数无效'}), 200

    events = Event.query.filter(Event.clubID == club_id, Event.eventID.in_(picked_ids)).all()
    if not events:
        return jsonify({'code': 4004, 'message': '未找到可导出的活动'}), 200

    event_by_id = {e.eventID: e for e in events}
    sorted_events = [event_by_id[eid] for eid in picked_ids if eid in event_by_id]

    headers = ['活动名称', '参与人员', '参加时间', '打卡时间', '人员发布的动态']
    data_rows = []
    moment_files_per_row = []
    avatar_file_per_row = []
    cover_file_per_row = []

    def _fmt(dt):
        if not dt:
            return ''
        return dt.strftime('%Y-%m-%d %H:%M')

    join_date_null_last = case((EventJoin.joinDate.is_(None), 1), else_=0)
    for ev in sorted_events:
        event_cover = ev.cover
        event_joins = (
            EventJoin.query.filter_by(eventID=ev.eventID, isDelete=False)
            .order_by(join_date_null_last.asc(), EventJoin.joinDate.desc(), EventJoin.joinID.desc())
            .all()
        )
        moments = Moment.query.filter_by(ref_event_ID=ev.eventID).order_by(
            Moment.createDate.desc(), Moment.momentID.desc()
        ).all()
        moments_by_user = defaultdict(list)
        for m in moments:
            if m.creatorID is not None:
                moments_by_user[m.creatorID].append(m)

        if not event_joins:
            data_rows.append([ev.title or '', '无参与人员', '', '', ''])
            moment_files_per_row.append([])
            cover_file_per_row.append(event_cover)
            avatar_file_per_row.append(None)
            continue

        for ej in event_joins:
            user = ej.user
            if not user:
                continue
            user_moments = moments_by_user.get(user.userID, [])
            moment_lines = []
            for m in user_moments:
                desc = (m.description or '').strip() or '（无文字）'
                t = _fmt(m.createDate)
                moment_lines.append(f'{desc}（{t}）' if t else desc)
            data_rows.append([
                ev.title or '',
                user.userName or '',
                _fmt(ej.joinDate),
                _fmt(ej.clockinDate) or '未打卡',
                '；'.join(moment_lines) if moment_lines else '无',
            ])
            moment_files_per_row.append(collect_moment_files_from_moments(user_moments))
            cover_file_per_row.append(event_cover)
            avatar_file_per_row.append(user.avatar)

    if not PILLOW_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装图片处理库，无法打包动态原图'}), 200

    try:
        export_resp = create_participation_export_archive(
            headers=headers,
            data_rows=data_rows,
            moment_files_per_row=moment_files_per_row,
            filename_prefix=f'club_{club_id}_event_participation',
            excel_basename='活动参与明细',
            sheet_title='活动参与明细',
            avatar_file_per_row=avatar_file_per_row,
            cover_file_per_row=cover_file_per_row,
            export_layout='event_participation',
        )
        export_payload = export_resp.get_json(silent=True) or {}
        if export_payload.get('code') != 200:
            return jsonify({
                'code': export_payload.get('code', 5000),
                'message': export_payload.get('message', '导出失败'),
            }), 200

        file_data = export_payload.get('data') or {}
        object_path = file_data.get('file_path')
        file_name = file_data.get('filename') or f"club_{club_id}_event_participation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        if not object_path:
            return jsonify({'code': 5000, 'message': '导出结果缺少文件路径'}), 200

        media_id = upload_minio_object_to_wecom_media(object_path=object_path, file_name=file_name)
        return jsonify({
            'code': 200,
            'message': '已生成企业微信文件素材（含动态原图压缩包）',
            'data': {
                'media_id': media_id,
                'filename': file_name,
                'archive_format': file_data.get('archive_format', 'zip'),
                'club_id': club_id,
                'selected_event_count': len(sorted_events),
                'moment_image_count': file_data.get('moment_image_count', 0),
            },
        }), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'生成企业微信会话文件失败: {str(e)}'}), 200

@bp.route('/show/all_club/users', methods=['GET'])
@jwt_required()
def show_all_club_users():
    """显示所有协会的会员情况"""
    # 权限检查
    has_permission, message = check_permission(statistics.show_all_club_users.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    activity_start_date = request.args.get('activity_start_date', '').strip()
    activity_end_date = request.args.get('activity_end_date', '').strip()

    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        activity_start_dt, activity_end_dt = parse_date_range(activity_start_date, activity_end_date)
        # 获取所有协会
        clubs = Club.query.all()
        selected_user_club_ids = {}
        result_data = []
        for club in clubs:
            # 获取该协会的所有成员
            members = ClubMember.query.filter(ClubMember.clubID == club.clubID).all()
            filtered_members = [member for member in members if is_member_in_range(member, start_dt, end_dt)]
            
            member_list = []
            for member in filtered_members:
                user = User.query.filter(User.userID == member.userID).first()
                if user:
                    if user.userID not in selected_user_club_ids:
                        selected_user_club_ids[user.userID] = set()
                    selected_user_club_ids[user.userID].add(club.clubID)
                    member_list.append({
                        'user_id': user.userID,
                        'user_name': user.userName,
                        'gender': user.gender or '',
                        'phone_number': user.phone or '',
                        'email': user.email or '',
                        'department': user.department or '',
                        'role': member.role or 'member',
                        'join_date': member.joinDate.strftime('%Y-%m-%d %H:%M:%S') if member.joinDate else '',
                        'create_date': user.createDate.strftime('%Y-%m-%d %H:%M:%S') if user.createDate else ''
                    })
            
            result_data.append({
                'club_id': club.clubID,
                'club_name': club.clubName,
                'member_count': len(member_list),
                'members': member_list
            })

        total_dynamic_count = 0
        for uid, club_ids in selected_user_club_ids.items():
            total_dynamic_count += count_user_moments_in_clubs(
                user_id=uid,
                club_ids=list(club_ids),
                start_date=activity_start_dt,
                end_date=activity_end_dt
            )
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'clubs': result_data,
                'total_clubs': len(result_data),
                'total_members': sum(club['member_count'] for club in result_data),
                'total_dynamic_count': total_dynamic_count,
                'start_date': start_dt.strftime('%Y-%m-%d') if start_dt else '',
                'end_date': end_dt.strftime('%Y-%m-%d') if end_dt else '',
                'activity_start_date': activity_start_dt.strftime('%Y-%m-%d') if activity_start_dt else '',
                'activity_end_date': activity_end_dt.strftime('%Y-%m-%d') if activity_end_dt else ''
            }
        })
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'获取数据失败: {str(e)}'}), 200

@bp.route('/show/club/<int:club_id>/all_event/details', methods=['GET'])
@jwt_required()
def show_club_all_event_details(club_id):
    """显示指定协会的所有活动详细信息"""
    # 权限检查
    has_permission, message = check_permission(statistics.show_club_all_event_details.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    try:
        # 验证协会是否存在
        club = Club.query.filter_by(clubID=club_id).first()
        if not club:
            return jsonify({'code': 4004, 'message': '协会不存在'}), 200

        start_dt, end_dt = parse_date_range(start_date, end_date)
        club_events = Event.query.filter_by(clubID=club_id).all()
        events = [event for event in club_events if is_event_in_range(event, start_dt, end_dt)]
        
        event_list = []
        for event in events:
            # 获取报名人数和签到人数
            total_participants = EventJoin.query.filter_by(eventID=event.eventID).count()
            checked_in_count = EventJoin.query.filter_by(eventID=event.eventID).filter(
                EventJoin.clockinDate.isnot(None)
            ).count()
            
            # 获取组织者信息
            organizer = User.query.filter_by(userID=event.authorID).first()
            
            event_list.append({
                'event_id': event.eventID,
                'title': event.title,
                'message': event.message or '',
                'location_name': event.location_name or event.location or '',
                'location_address': event.location_address or '',
                'pre_start_time': event.pre_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_startTime else '',
                'pre_end_time': event.pre_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_endTime else '',
                'actual_start_time': event.actual_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_startTime else '',
                'actual_end_time': event.actual_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_endTime else '',
                'budget': event.budget or 0,
                'real_cost': event.real_cost or 0,
                'total_participants': total_participants,
                'checked_in_count': checked_in_count,
                'create_time': event.createDate.strftime('%Y-%m-%d %H:%M:%S') if event.createDate else '',
                'cover_url': event.cover.fileUrl if event.cover else None,
                'is_cancelled': bool(event.is_cancelled),
                'club_deleted': bool(event.club.isDelete) if event.club else False,
                'organizer': {
                    'user_id': organizer.userID if organizer else None,
                    'user_name': organizer.userName if organizer else '未知'
                }
            })
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'club': {
                    'club_id': club.clubID,
                    'club_name': club.clubName
                },
                'events': event_list,
                'total_events': len(event_list),
                'total_participants': sum(event['total_participants'] for event in event_list),
                'total_checked_in': sum(event['checked_in_count'] for event in event_list),
                'start_date': start_dt.strftime('%Y-%m-%d') if start_dt else '',
                'end_date': end_dt.strftime('%Y-%m-%d') if end_dt else ''
            }
        })

    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'获取数据失败: {str(e)}'}), 200

@bp.route('/show/all_club/all_event/details', methods=['GET'])
@jwt_required()
def show_all_club_all_event_details():
    """显示所有协会的所有活动详细信息"""
    # 权限检查
    has_permission, message = check_permission(statistics.show_all_club_all_event_details.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        all_events = Event.query.join(Club).all()
        events = [event for event in all_events if is_event_in_range(event, start_dt, end_dt)]
        
        event_list = []
        for event in events:
            # 获取报名人数和签到人数
            total_participants = EventJoin.query.filter_by(eventID=event.eventID).count()
            checked_in_count = EventJoin.query.filter_by(eventID=event.eventID).filter(
                EventJoin.clockinDate.isnot(None)
            ).count()
            
            # 获取组织者信息
            organizer = User.query.filter_by(userID=event.authorID).first()
            
            event_list.append({
                'event_id': event.eventID,
                'title': event.title,
                'message': event.message or '',
                'club': {
                    'club_id': event.club.clubID if event.club else None,
                    'club_name': event.club.clubName if event.club else '未知协会'
                },
                'location_name': event.location_name or event.location or '',
                'location_address': event.location_address or '',
                'pre_start_time': event.pre_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_startTime else '',
                'pre_end_time': event.pre_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_endTime else '',
                'actual_start_time': event.actual_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_startTime else '',
                'actual_end_time': event.actual_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_endTime else '',
                'budget': event.budget or 0,
                'real_cost': event.real_cost or 0,
                'total_participants': total_participants,
                'checked_in_count': checked_in_count,
                'create_time': event.createDate.strftime('%Y-%m-%d %H:%M:%S') if event.createDate else '',
                'organizer': {
                    'user_id': organizer.userID if organizer else None,
                    'user_name': organizer.userName if organizer else '未知'
                }
            })
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'events': event_list,
                'total_events': len(event_list),
                'total_participants': sum(event['total_participants'] for event in event_list),
                'total_checked_in': sum(event['checked_in_count'] for event in event_list)
            }
        })
        
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'获取数据失败: {str(e)}'}), 200

@bp.route('/show/event/<int:event_id>/details', methods=['GET'])
@jwt_required()
def show_event_details(event_id):
    """显示指定活动的详细信息（包含参与者列表）"""
    # 权限检查
    has_permission, message = check_permission(statistics.show_event_details.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    # 验证活动是否存在
    event = Event.query.filter_by(eventID=event_id).first()
    if not event:
        return jsonify({'code': 4004, 'message': '活动不存在'}), 200
    
    try:
        # 获取活动参与者
        event_joins = EventJoin.query.filter_by(eventID=event_id).all()
        
        # 获取组织者信息
        organizer = User.query.filter_by(userID=event.authorID).first()
        
        participant_list = []
        for join in event_joins:
            user = User.query.filter_by(userID=join.userID).first()
            if user:
                participant_list.append({
                    'user_id': user.userID,
                    'user_name': user.userName,
                    'gender': user.gender or '',
                    'phone_number': user.phone or '',
                    'email': user.email or '',
                    'department': user.department or '',
                    'join_date': join.createDate.strftime('%Y-%m-%d %H:%M:%S') if join.createDate else '',
                    'checkin_date': join.clockinDate.strftime('%Y-%m-%d %H:%M:%S') if join.clockinDate else '',
                    'is_checked_in': bool(join.clockinDate),
                    'note': join.note or ''
                })
        
        # 统计信息
        total_participants = len(participant_list)
        checked_in_count = sum(1 for p in participant_list if p['is_checked_in'])
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'event': {
                    'event_id': event.eventID,
                    'title': event.title,
                    'message': event.message or '',
                    'club': {
                        'club_id': event.club.clubID if event.club else None,
                        'club_name': event.club.clubName if event.club else '未知协会'
                    },
                    'location_name': event.location_name or event.location or '',
                    'location_address': event.location_address or '',
                    'pre_start_time': event.pre_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_startTime else '',
                    'pre_end_time': event.pre_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.pre_endTime else '',
                    'actual_start_time': event.actual_startTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_startTime else '',
                    'actual_end_time': event.actual_endTime.strftime('%Y-%m-%d %H:%M:%S') if event.actual_endTime else '',
                    'budget': event.budget or 0,
                    'real_cost': event.real_cost or 0,
                    'create_time': event.createDate.strftime('%Y-%m-%d %H:%M:%S') if event.createDate else '',
                    'organizer': {
                        'user_id': organizer.userID if organizer else None,
                        'user_name': organizer.userName if organizer else '未知'
                    }
                },
                'participants': participant_list,
                'statistics': {
                    'total_participants': total_participants,
                    'checked_in_count': checked_in_count,
                    'checkin_rate': round(checked_in_count / total_participants * 100, 2) if total_participants > 0 else 0
                }
            }
        })
        
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'获取数据失败: {str(e)}'}), 200

@bp.route('/show/club/<int:club_id>/financial/statistics', methods=['GET'])
@jwt_required()
def show_club_financial_statistics(club_id):
    """显示指定协会的收支统计信息"""
    # 权限检查
    has_permission, message = check_permission(statistics.show_club_financial_statistics.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    # 验证协会是否存在
    club = Club.query.filter_by(clubID=club_id).first()
    if not club:
        return jsonify({'code': 4004, 'message': '协会不存在'}), 200
    
    # 获取时间参数
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    try:
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            # 将结束日期设为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # 默认显示当年的数据
            current_year = datetime.now().year
            start_date = datetime(current_year, 1, 1)
            end_date = datetime(current_year, 12, 31, 23, 59, 59)
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    
    try:
        # 1. 计算支出
        # 1.1 ClubFee支出
        club_fees = ClubFee.query.filter(
            ClubFee.clubID == club_id,
            ClubFee.createDate >= start_date,
            ClubFee.createDate <= end_date
        ).all()
        
        club_fee_total = sum(fee.feement for fee in club_fees)
        club_fee_details = [{
            'fee_id': fee.feeID,
            'amount': fee.feement,
            'description': fee.description,
            'create_date': fee.createDate.strftime('%Y-%m-%d %H:%M:%S') if fee.createDate else ''
        } for fee in club_fees]
        
        # 1.2 活动实际费用支出
        events = Event.query.filter(Event.clubID == club_id).all()
        event_cost_total = 0
        event_cost_details = []
        
        for event in events:
            event_date = event.actual_startTime if event.actual_startTime else event.pre_startTime
            if event_date and start_date <= event_date <= end_date and event.real_cost:
                event_cost_total += event.real_cost
                event_cost_details.append({
                    'event_id': event.eventID,
                    'title': event.title,
                    'amount': event.real_cost,
                    'date': event_date.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        total_expenses = club_fee_total + event_cost_total
        
        # 2. 计算收入
        # 2.1 已完成的PayGroup收入（与该协会相关的）
        completed_pay_groups = PayGroup.query.filter(
            PayGroup.clubID == club_id,
            PayGroup.createDate >= start_date,
            PayGroup.createDate <= end_date
        ).all()
        
        paygroup_income_total = 0
        paygroup_income_details = []
        
        for pay_group in completed_pay_groups:
            # 计算已收到的款项
            paid_amount = sum(p.payment for p in pay_group.paypersonals if p.payDate)
            paygroup_income_total += paid_amount
            paygroup_income_details.append({
                'group_id': pay_group.groupID,
                'description': pay_group.description,
                'total_payment': pay_group.totalpayment,
                'paid_amount': paid_amount,
                'unpaid_amount': pay_group.totalpayment - paid_amount,
                'create_date': pay_group.createDate.strftime('%Y-%m-%d %H:%M:%S') if pay_group.createDate else '',
                'participants_count': len(pay_group.paypersonals),
                'paid_participants': len([p for p in pay_group.paypersonals if p.payDate])
            })
        
        total_income = paygroup_income_total
        
        # 3. 计算净收支
        net_balance = total_income - total_expenses
        
        # 4. 按月份统计
        monthly_stats = {}
        
        # 按月份统计支出
        for fee in club_fees:
            month_key = fee.createDate.strftime('%Y-%m') if fee.createDate else 'unknown'
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'income': 0, 'expenses': 0}
            monthly_stats[month_key]['expenses'] += fee.feement
        
        for event in events:
            event_date = event.actual_startTime if event.actual_startTime else event.pre_startTime
            if event_date and start_date <= event_date <= end_date and event.real_cost:
                month_key = event_date.strftime('%Y-%m')
                if month_key not in monthly_stats:
                    monthly_stats[month_key] = {'income': 0, 'expenses': 0}
                monthly_stats[month_key]['expenses'] += event.real_cost
        
        # 按月份统计收入
        for pay_group in completed_pay_groups:
            month_key = pay_group.createDate.strftime('%Y-%m') if pay_group.createDate else 'unknown'
            paid_amount = sum(p.payment for p in pay_group.paypersonals if p.payDate)
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'income': 0, 'expenses': 0}
            monthly_stats[month_key]['income'] += paid_amount
        
        # 转换月度统计为列表格式
        monthly_list = []
        for month_key in sorted(monthly_stats.keys()):
            if month_key != 'unknown':
                monthly_data = monthly_stats[month_key]
                monthly_list.append({
                    'month': month_key,
                    'income': monthly_data['income'],
                    'expenses': monthly_data['expenses'],
                    'net_balance': monthly_data['income'] - monthly_data['expenses']
                })
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'club': {
                    'club_id': club.clubID,
                    'club_name': club.clubName
                },
                'time_range': {
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d')
                },
                'summary': {
                    'total_income': total_income,
                    'total_expenses': total_expenses,
                    'net_balance': net_balance,
                    'club_fee_expenses': club_fee_total,
                    'event_cost_expenses': event_cost_total,
                    'paygroup_income': paygroup_income_total
                },
                'income_details': {
                    'paygroups': paygroup_income_details
                },
                'expense_details': {
                    'club_fees': club_fee_details,
                    'event_costs': event_cost_details
                },
                'monthly_statistics': monthly_list
            }
        })
        
    except Exception as e:
        return jsonify({'code': 5000, 'message': f'获取数据失败: {str(e)}'}), 200

@bp.route('/export/club/<int:club_id>/financial/statistics', methods=['GET'])
@jwt_required()
def export_club_financial_statistics(club_id):
    """导出指定协会的收支统计信息到Excel"""
    # 权限检查
    has_permission, message = check_permission(statistics.export_club_financial_statistics.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    # 验证协会是否存在
    club = Club.query.filter_by(clubID=club_id).first()
    if not club:
        return jsonify({'code': 4004, 'message': '协会不存在'}), 200
    
    if not EXCEL_AVAILABLE:
        return jsonify({'code': 5000, 'message': '服务器未安装Excel支持库'}), 200
    
    # 获取时间参数
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    try:
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            # 将结束日期设为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # 默认导出当年的数据
            current_year = datetime.now().year
            start_date = datetime(current_year, 1, 1)
            end_date = datetime(current_year, 12, 31, 23, 59, 59)
    except ValueError:
        return jsonify({'code': 4001, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 200
    
    try:
        # 获取统计数据（复用上面的逻辑）
        # 1. 支出数据
        club_fees = ClubFee.query.filter(
            ClubFee.clubID == club_id,
            ClubFee.createDate >= start_date,
            ClubFee.createDate <= end_date
        ).all()
        
        events = Event.query.filter(Event.clubID == club_id).all()
        event_costs = []
        for event in events:
            event_date = event.actual_startTime if event.actual_startTime else event.pre_startTime
            if event_date and start_date <= event_date <= end_date and event.real_cost:
                event_costs.append(event)
        
        # 2. 收入数据
        completed_pay_groups = PayGroup.query.filter(
            PayGroup.clubID == club_id,
            PayGroup.createDate >= start_date,
            PayGroup.createDate <= end_date
        ).all()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 工作表1：汇总统计
        ws_summary = wb.active
        ws_summary.title = "收支汇总"
        
        # 计算汇总数据
        club_fee_total = sum(fee.feement for fee in club_fees)
        event_cost_total = sum(event.real_cost for event in event_costs)
        total_expenses = club_fee_total + event_cost_total
        
        paygroup_income_total = sum(
            sum(p.payment for p in pg.paypersonals if p.payDate) 
            for pg in completed_pay_groups
        )
        total_income = paygroup_income_total
        net_balance = total_income - total_expenses
        
        # 写入汇总数据
        summary_data = [
            ['协会名称', club.clubName],
            ['统计时间范围', f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}"],
            ['', ''],
            ['收入项目', '金额(元)'],
            ['缴费收入', paygroup_income_total],
            ['收入小计', total_income],
            ['', ''],
            ['支出项目', '金额(元)'],
            ['协会费用', club_fee_total],
            ['活动费用', event_cost_total],
            ['支出小计', total_expenses],
            ['', ''],
            ['净收支', net_balance]
        ]
        
        for row_idx, (item, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=item)
            ws_summary.cell(row=row_idx, column=2, value=value)
            
            # 设置样式
            if item in ['收入项目', '支出项目'] or row_idx == 1:
                ws_summary.cell(row=row_idx, column=1).font = header_font
                ws_summary.cell(row=row_idx, column=1).fill = header_fill
                ws_summary.cell(row=row_idx, column=2).font = header_font
                ws_summary.cell(row=row_idx, column=2).fill = header_fill
        
        # 设置列宽
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 15
        
        # 工作表2：支出明细
        ws_expenses = wb.create_sheet("支出明细")
        expense_headers = ['类型', '项目ID', '项目名称', '金额(元)', '日期', '描述']
        
        for col, header in enumerate(expense_headers, 1):
            cell = ws_expenses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        current_row = 2
        
        # 协会费用明细
        for fee in club_fees:
            ws_expenses.cell(row=current_row, column=1, value='协会费用')
            ws_expenses.cell(row=current_row, column=2, value=fee.feeID)
            ws_expenses.cell(row=current_row, column=3, value='协会日常支出')
            ws_expenses.cell(row=current_row, column=4, value=fee.feement)
            ws_expenses.cell(row=current_row, column=5, value=fee.createDate.strftime('%Y-%m-%d') if fee.createDate else '')
            ws_expenses.cell(row=current_row, column=6, value=fee.description or '')
            current_row += 1
        
        # 活动费用明细
        for event in event_costs:
            event_date = event.actual_startTime if event.actual_startTime else event.pre_startTime
            ws_expenses.cell(row=current_row, column=1, value='活动费用')
            ws_expenses.cell(row=current_row, column=2, value=event.eventID)
            ws_expenses.cell(row=current_row, column=3, value=event.title)
            ws_expenses.cell(row=current_row, column=4, value=event.real_cost)
            ws_expenses.cell(row=current_row, column=5, value=event_date.strftime('%Y-%m-%d') if event_date else '')
            ws_expenses.cell(row=current_row, column=6, value='活动实际费用')
            current_row += 1
        
        # 设置列宽
        for col_letter, width in [('A', 12), ('B', 10), ('C', 25), ('D', 12), ('E', 12), ('F', 30)]:
            ws_expenses.column_dimensions[col_letter].width = width
        
        # 工作表3：收入明细
        ws_income = wb.create_sheet("收入明细")
        income_headers = ['缴费组ID', '描述', '应收金额(元)', '已收金额(元)', '未收金额(元)', '创建日期', '参与人数', '已缴费人数']
        
        for col, header in enumerate(income_headers, 1):
            cell = ws_income.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        current_row = 2
        for pay_group in completed_pay_groups:
            paid_amount = sum(p.payment for p in pay_group.paypersonals if p.payDate)
            ws_income.cell(row=current_row, column=1, value=pay_group.groupID)
            ws_income.cell(row=current_row, column=2, value=pay_group.description or '')
            ws_income.cell(row=current_row, column=3, value=pay_group.totalpayment)
            ws_income.cell(row=current_row, column=4, value=paid_amount)
            ws_income.cell(row=current_row, column=5, value=pay_group.totalpayment - paid_amount)
            ws_income.cell(row=current_row, column=6, value=pay_group.createDate.strftime('%Y-%m-%d') if pay_group.createDate else '')
            ws_income.cell(row=current_row, column=7, value=len(pay_group.paypersonals))
            ws_income.cell(row=current_row, column=8, value=len([p for p in pay_group.paypersonals if p.payDate]))
            current_row += 1
        
        # 设置列宽
        for col_letter, width in [('A', 12), ('B', 25), ('C', 15), ('D', 15), ('E', 15), ('F', 12), ('G', 12), ('H', 12)]:
            ws_income.column_dimensions[col_letter].width = width
        
        # 冻结首行
        ws_summary.freeze_panes = 'A2'
        ws_expenses.freeze_panes = 'A2'
        ws_income.freeze_panes = 'A2'
        
        # 保存到内存
        output = io.BytesIO()
        try:
            wb.save(output)
            output.seek(0)
            
            # 上传到MinIO
            minio_client = get_minio_client()
            bucket_name = current_app.config.get('MINIO_BUCKET', 'manage-mate')
            ensure_bucket_exists(minio_client, bucket_name)
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"club_{club_id}_financial_statistics_{timestamp}.xlsx"
            file_path = f"statistics/{filename}"
            
            # 获取文件大小
            file_size = output.getbuffer().nbytes
            
            minio_client.put_object(
                bucket_name,
                file_path,
                output,
                length=file_size,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # 在上传完成后关闭工作簿，释放所有资源
            wb.close()
            
        finally:
            # 确保关闭输出流
            if output:
                try:
                    output.close()
                except:
                    pass
        
        # 生成下载URL
        base_url = (current_app.config.get('BASE_URL') or 'https://www.vhhg.top').rstrip('/')
        download_url = f"{base_url}/api/v1/file/download/tmp/{file_path}"
        
        current_app.logger.info(f"协会{club_id}收支统计Excel文件生成并上传成功: {file_path}")
        
        return jsonify({
            'code': 200,
            'message': '导出成功',
            'data': {
                'download_url': download_url,
                'filename': filename,
                'file_path': file_path,
                'file_size': file_size,
                'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'club_name': club.clubName,
                'time_range': {
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d')
                }
            }
        })
        
    except S3Error as e:
        current_app.logger.error(f"MinIO上传失败: {str(e)}")
        return jsonify({'code': 5000, 'message': f'文件上传失败: {str(e)}'}), 200
    except Exception as e:
        current_app.logger.error(f'导出协会{club_id}收支统计失败: {str(e)}')
        return jsonify({'code': 5000, 'message': f'导出失败: {str(e)}'}), 200

def collect_moment_files_from_moments(moments):
    """从动态列表按 imageIDs 顺序收集图片文件对象。"""
    from app.models.file import File

    image_ids = []
    for moment in moments or []:
        if not moment.imageIDs or not isinstance(moment.imageIDs, list):
            continue
        for image_id in moment.imageIDs:
            if image_id not in image_ids:
                image_ids.append(image_id)
    if not image_ids:
        return []
    files = File.query.filter(File.fileID.in_(image_ids)).all()
    file_map = {f.fileID: f for f in files}
    return [file_map[fid] for fid in image_ids if fid in file_map]


def _participation_row_asset_folder_name(row_idx, row_data):
    """根据前两列文本生成动态原图子文件夹名。"""
    skip_values = {'无活动记录', '无参与人员', '无', ''}
    parts = []
    for value in (row_data or [])[:2]:
        text = str(value or '').strip()
        if not text or text in skip_values:
            continue
        parts.append(sanitize_filename(text))
    if not parts:
        parts = [f'行{row_idx:03d}']
    return f"{'-'.join(parts[:2])}-资源"


def _write_participation_single_image_cell(
    ws,
    row_idx,
    col_idx,
    file_obj,
    row_asset_folder,
    excel_path,
    temp_root,
    temp_thumb_files,
    file_basename,
    thumb_prefix,
    empty_label,
    link_label,
    minlength,
):
    """向 Excel 单元格写入单张图片略缩图，并把原图落到行资源目录。"""
    cell = ws.cell(row=row_idx, column=col_idx, value=empty_label)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if not file_obj or not getattr(file_obj, 'fileUrl', None):
        return 0
    image_bytes = download_image_from_minio(file_obj.fileUrl)
    if not image_bytes:
        return 0
    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
    image_name = f'{file_basename}{ext or ".jpg"}'
    image_abs = os.path.join(row_asset_folder, image_name)
    with open(image_abs, 'wb') as f:
        f.write(image_bytes)

    thumb_bytes = generate_thumbnail_bytes(image_bytes, minlength=minlength)
    if thumb_bytes:
        fd, thumb_path = tempfile.mkstemp(prefix=thumb_prefix, suffix='.jpg', dir=temp_root)
        os.close(fd)
        with open(thumb_path, 'wb') as f:
            f.write(thumb_bytes)
        temp_thumb_files.append(thumb_path)
        col_letter = get_excel_column_name(col_idx)
        excel_img = ExcelImage(thumb_path)
        excel_img.anchor = f'{col_letter}{row_idx}'
        ws.add_image(excel_img)

    relative_link = os.path.relpath(image_abs, os.path.dirname(excel_path)).replace('\\', '/')
    cell.value = link_label
    cell.hyperlink = relative_link
    cell.style = 'Hyperlink'
    return 1


def create_participation_export_archive(
    headers,
    data_rows,
    moment_files_per_row,
    filename_prefix,
    excel_basename='参与明细',
    sheet_title='参与明细',
    avatar_file_per_row=None,
    cover_file_per_row=None,
    export_layout='basic',
):
    """参与/成员活动明细 ZIP 导出：Excel（封面/头像/动态略缩图+超链接）+ 原图文件夹。"""
    if not EXCEL_AVAILABLE:
        raise Exception('Excel支持库未安装')
    if not PILLOW_AVAILABLE:
        raise Exception('图片处理库未安装')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    package_name = f'{filename_prefix}_{timestamp}'
    temp_root = tempfile.mkdtemp(prefix='participation_archive_')
    package_root = os.path.join(temp_root, package_name)
    os.makedirs(package_root, exist_ok=True)

    excel_path = os.path.join(package_root, f'{sanitize_filename(excel_basename)}.xlsx')
    asset_root = os.path.join(package_root, '导出资源')
    os.makedirs(asset_root, exist_ok=True)

    row_count = len(data_rows)
    if len(moment_files_per_row) < row_count:
        moment_files_per_row = list(moment_files_per_row) + [[]] * (row_count - len(moment_files_per_row))
    elif len(moment_files_per_row) > row_count:
        moment_files_per_row = moment_files_per_row[:row_count]

    if avatar_file_per_row is None:
        avatar_file_per_row = [None] * row_count
    elif len(avatar_file_per_row) < row_count:
        avatar_file_per_row = list(avatar_file_per_row) + [None] * (row_count - len(avatar_file_per_row))
    elif len(avatar_file_per_row) > row_count:
        avatar_file_per_row = avatar_file_per_row[:row_count]

    if cover_file_per_row is None:
        cover_file_per_row = [None] * row_count
    elif len(cover_file_per_row) < row_count:
        cover_file_per_row = list(cover_file_per_row) + [None] * (row_count - len(cover_file_per_row))
    elif len(cover_file_per_row) > row_count:
        cover_file_per_row = cover_file_per_row[:row_count]

    if export_layout == 'member_activity':
        full_headers = [
            '用户头像略缩图', '用户姓名', '活动封面略缩图', '活动名称',
            '参加时间', '打卡时间', '人员发布的动态', '动态图片略缩图',
        ]
    elif export_layout == 'event_participation':
        full_headers = [
            '活动封面略缩图', '活动名称', '用户头像略缩图', '参与人员',
            '参加时间', '打卡时间', '人员发布的动态', '动态图片略缩图',
        ]
    else:
        full_headers = list(headers) + ['动态图片略缩图']

    temp_thumb_files = []
    moment_image_count = 0
    cover_image_count = 0
    avatar_image_count = 0
    used_folder_names = set()
    moment_col = len(full_headers)
    cover_col = None
    avatar_col = None
    if export_layout == 'member_activity':
        avatar_col, cover_col = 1, 3
    elif export_layout == 'event_participation':
        cover_col, avatar_col = 1, 3

    wb = None
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sanitize_sheet_title(sheet_title)

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for col, header in enumerate(full_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        ws.row_dimensions[1].height = 25

        moment_col_letter = get_excel_column_name(moment_col)
        for row_idx, row_data in enumerate(data_rows, 2):
            ws.row_dimensions[row_idx].height = 90
            row_asset_folder = None

            def ensure_row_asset_folder():
                nonlocal row_asset_folder
                if row_asset_folder:
                    return row_asset_folder
                folder_base = _participation_row_asset_folder_name(row_idx, row_data)
                folder_name = folder_base
                suffix = 1
                while folder_name in used_folder_names:
                    suffix += 1
                    folder_name = f'{folder_base}_{suffix}'
                used_folder_names.add(folder_name)
                row_asset_folder = os.path.join(asset_root, folder_name)
                os.makedirs(row_asset_folder, exist_ok=True)
                return row_asset_folder

            if export_layout == 'member_activity':
                for col_offset, value in enumerate(row_data, 2):
                    if col_offset in (avatar_col, cover_col):
                        continue
                    cell = ws.cell(row=row_idx, column=col_offset, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                avatar_file = avatar_file_per_row[row_idx - 2]
                cover_file = cover_file_per_row[row_idx - 2]
                if avatar_file:
                    avatar_image_count += _write_participation_single_image_cell(
                        ws, row_idx, avatar_col, avatar_file, ensure_row_asset_folder(),
                        excel_path, temp_root, temp_thumb_files,
                        'avatar', 'part_avatar_thumb_', '无头像', '头像原图', 30,
                    )
                else:
                    ws.cell(row=row_idx, column=avatar_col, value='无头像').alignment = center_alignment
                if cover_file:
                    cover_image_count += _write_participation_single_image_cell(
                        ws, row_idx, cover_col, cover_file, ensure_row_asset_folder(),
                        excel_path, temp_root, temp_thumb_files,
                        'cover', 'part_cover_thumb_', '无封面', '封面原图', 50,
                    )
                else:
                    ws.cell(row=row_idx, column=cover_col, value='无封面').alignment = center_alignment
            elif export_layout == 'event_participation':
                for col_offset, value in enumerate(row_data, 2):
                    if col_offset in (avatar_col, cover_col):
                        continue
                    cell = ws.cell(row=row_idx, column=col_offset, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                cover_file = cover_file_per_row[row_idx - 2]
                avatar_file = avatar_file_per_row[row_idx - 2]
                if cover_file:
                    cover_image_count += _write_participation_single_image_cell(
                        ws, row_idx, cover_col, cover_file, ensure_row_asset_folder(),
                        excel_path, temp_root, temp_thumb_files,
                        'cover', 'part_cover_thumb_', '无封面', '封面原图', 50,
                    )
                else:
                    ws.cell(row=row_idx, column=cover_col, value='无封面').alignment = center_alignment
                if avatar_file:
                    avatar_image_count += _write_participation_single_image_cell(
                        ws, row_idx, avatar_col, avatar_file, ensure_row_asset_folder(),
                        excel_path, temp_root, temp_thumb_files,
                        'avatar', 'part_avatar_thumb_', '无头像', '头像原图', 30,
                    )
                else:
                    ws.cell(row=row_idx, column=avatar_col, value='无头像').alignment = center_alignment
            else:
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment

            img_cell = ws.cell(row=row_idx, column=moment_col, value='无动态图片')
            img_cell.font = data_font
            img_cell.alignment = center_alignment
            moment_files = moment_files_per_row[row_idx - 2] if row_idx - 2 < len(moment_files_per_row) else []

            if moment_files:
                row_folder = ensure_row_asset_folder()
                saved_paths = []
                for idx, file_obj in enumerate(moment_files, 1):
                    if not file_obj.fileUrl:
                        continue
                    image_bytes = download_image_from_minio(file_obj.fileUrl)
                    if not image_bytes:
                        continue
                    moment_image_count += 1
                    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
                    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
                    image_name = f'moment_{idx}{ext or ".jpg"}'
                    image_abs = os.path.join(row_folder, image_name)
                    with open(image_abs, 'wb') as f:
                        f.write(image_bytes)
                    saved_paths.append(image_abs)

                collage_bytes = build_moment_collage_thumbnail(moment_files, minlength=50, max_images=12)
                if collage_bytes:
                    fd, collage_path = tempfile.mkstemp(prefix='part_moment_thumb_', suffix='.jpg', dir=temp_root)
                    os.close(fd)
                    with open(collage_path, 'wb') as f:
                        f.write(collage_bytes)
                    temp_thumb_files.append(collage_path)
                    excel_collage = ExcelImage(collage_path)
                    excel_collage.anchor = f'{moment_col_letter}{row_idx}'
                    ws.add_image(excel_collage)
                    img_cell.value = ''

                if saved_paths:
                    relative_link = os.path.relpath(saved_paths[0], os.path.dirname(excel_path)).replace('\\', '/')
                    img_cell.value = '动态原图'
                    img_cell.hyperlink = relative_link
                    img_cell.style = 'Hyperlink'

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value or ''))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 8), 36)
        if export_layout in ('member_activity', 'event_participation'):
            ws.column_dimensions[get_excel_column_name(1)].width = 18
            ws.column_dimensions[get_excel_column_name(3)].width = 18
        ws.column_dimensions[moment_col_letter].width = 28
        ws.freeze_panes = 'A2'
        wb.save(excel_path)

        archive_info = create_export_archive(
            source_dir=package_root,
            output_dir=temp_root,
            package_name=package_name,
        )
        minio_response = upload_local_file_to_minio(
            local_file_path=archive_info['archive_path'],
            object_path=f"statistics/{archive_info['archive_filename']}",
            content_type=archive_info['content_type'],
        )

        return jsonify({
            'code': 200,
            'message': archive_info['message'],
            'data': {
                **minio_response,
                'archive_format': archive_info['archive_format'],
                'row_count': len(data_rows),
                'moment_image_count': moment_image_count,
                'cover_image_count': cover_image_count,
                'avatar_image_count': avatar_image_count,
            },
        })
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        for tmp_file in temp_thumb_files:
            if tmp_file and os.path.exists(tmp_file):
                try:
                    os.unlink(tmp_file)
                except Exception:
                    pass
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass


def get_excel_column_name(col_num):
    """
    将列号转换为Excel列名
    例如: 1->A, 2->B, 26->Z, 27->AA, 28->AB, 702->ZZ, 703->AAA
    """
    result = ""
    while col_num > 0:
        col_num -= 1  # 转换为0-based索引
        result = chr(ord('A') + col_num % 26) + result
        col_num //= 26
    return result

def parse_date_range(start_date_str, end_date_str):
    """解析日期区间。两个参数都为空时表示不筛选。"""
    if not start_date_str and not end_date_str:
        return None, None

    if not start_date_str or not end_date_str:
        raise ValueError("start_date 和 end_date 必须同时传入")

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    if start_date > end_date:
        raise ValueError("开始日期不能大于结束日期")
    return start_date, end_date

def get_event_reference_datetime(event):
    """活动时间优先级：预开始 > 实际开始 > 创建时间。"""
    return event.pre_startTime or event.actual_startTime or event.createDate

def is_event_in_range(event, start_date, end_date):
    """判断活动是否在时间范围内。"""
    if not start_date or not end_date:
        return True
    ref_dt = get_event_reference_datetime(event)
    if not ref_dt:
        return False
    return start_date <= ref_dt <= end_date

def is_member_in_range(member, start_date, end_date):
    """判断会员加入记录是否在时间范围内。"""
    if not start_date or not end_date:
        return True
    ref_dt = member.joinDate
    if not ref_dt:
        return False
    return start_date <= ref_dt <= end_date

def sanitize_sheet_title(sheet_title):
    """Excel sheet 名称清洗并截断到31字符。"""
    safe_title = re.sub(r'[\\/*?:\[\]]', '_', (sheet_title or '未命名协会')).strip()
    if not safe_title:
        safe_title = '未命名协会'
    return safe_title[:31]

def sanitize_filename(file_name):
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', (file_name or '').strip())
    return safe_name or '未命名'

def extract_file_path_from_download_url(file_url):
    """从对外 fileUrl 解析 MinIO 对象路径；仅接受当前 BASE_URL 或站内相对路径。"""
    if not file_url:
        return None
    file_url = file_url.strip()
    base_url = (current_app.config.get('BASE_URL') or 'https://www.vhhg.top').rstrip('/')
    abs_prefix = f"{base_url}/api/v1/file/download/"
    rel_prefix = '/api/v1/file/download/'
    tail = None
    if file_url.startswith(abs_prefix):
        tail = file_url[len(abs_prefix):]
    elif file_url.startswith(rel_prefix):
        tail = file_url[len(rel_prefix):]
    if not tail:
        return None
    if '?' in tail:
        tail = tail.split('?', 1)[0]
    tail = tail.strip()
    return tail or None

def generate_thumbnail_bytes(image_data, minlength=50):
    """参考 download_thumbnail 生成略缩图。"""
    if not PILLOW_AVAILABLE:
        return None

    try:
        with PILImage.open(io.BytesIO(image_data)) as img:
            width, height = img.size
            if width <= 0 or height <= 0:
                return None

            aspect_ratio = width / height
            if 0.8 <= aspect_ratio <= 1.25:
                target_width, target_height = minlength, minlength
            elif width > height:
                target_width, target_height = 2 * minlength, minlength
            else:
                target_width, target_height = minlength, 2 * minlength

            scale = max(target_width / width, target_height / height)
            scaled_width = int(width * scale)
            scaled_height = int(height * scale)
            resized = img.resize((scaled_width, scaled_height), PILImage.Resampling.LANCZOS)

            left = (scaled_width - target_width) // 2
            top = (scaled_height - target_height) // 2
            thumbnail = resized.crop((left, top, left + target_width, top + target_height))

            output = io.BytesIO()
            if img.mode in ('RGBA', 'LA', 'P'):
                converted = thumbnail.convert('RGB')
                converted.save(output, format='JPEG', quality=85)
            else:
                fmt = img.format if img.format in ('JPEG', 'PNG', 'GIF') else 'JPEG'
                if fmt == 'JPEG':
                    thumbnail.convert('RGB').save(output, format='JPEG', quality=85)
                else:
                    thumbnail.save(output, format=fmt)
            output.seek(0)
            return output.getvalue()
    except Exception as e:
        current_app.logger.warning(f"生成略缩图失败: {str(e)}")
        return None

def build_moment_collage_thumbnail(moment_files, minlength=50, max_images=6):
    """将多张动态略缩图拼接成单张图片，放入一个Excel单元格。"""
    if not PILLOW_AVAILABLE or not moment_files:
        return None

    thumbnails = []
    for file_obj in moment_files[:max_images]:
        img_data = download_image_from_minio(file_obj.fileUrl)
        if not img_data:
            continue
        thumb_data = generate_thumbnail_bytes(img_data, minlength=minlength)
        if not thumb_data:
            continue
        try:
            thumb = PILImage.open(io.BytesIO(thumb_data)).convert('RGB')
            thumbnails.append(thumb)
        except Exception:
            continue

    if not thumbnails:
        return None

    gap = 8
    total_width = sum(img.width for img in thumbnails) + gap * (len(thumbnails) - 1)
    max_height = max(img.height for img in thumbnails)
    canvas = PILImage.new('RGB', (total_width, max_height), (255, 255, 255))

    cursor_x = 0
    for img in thumbnails:
        y = (max_height - img.height) // 2
        canvas.paste(img, (cursor_x, y))
        cursor_x += img.width + gap

    output = io.BytesIO()
    canvas.save(output, format='JPEG', quality=85)
    output.seek(0)
    return output.getvalue()

def collect_event_moment_files(event):
    """按动态发布时间汇总活动动态图片文件。"""
    from app.models.file import File

    moments = Moment.query.filter_by(ref_event_ID=event.eventID).order_by(Moment.createDate.asc()).all()
    image_ids = []
    for moment in moments:
        if moment.imageIDs and isinstance(moment.imageIDs, list):
            for image_id in moment.imageIDs:
                if image_id not in image_ids:
                    image_ids.append(image_id)

    if not image_ids:
        return []

    files = File.query.filter(File.fileID.in_(image_ids)).all()
    file_map = {f.fileID: f for f in files}
    return [file_map[file_id] for file_id in image_ids if file_id in file_map]

def create_all_club_event_details_rar_package(events, filename_prefix, start_date=None, end_date=None):
    """每个协会单独一个Excel与原图文件夹，优先RAR，失败则ZIP。"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel支持库未安装")
    if not PILLOW_AVAILABLE:
        raise Exception("图片处理库未安装")

    grouped_events = defaultdict(list)
    for event in events:
        club_name = event.club.clubName if event.club else '未知协会'
        grouped_events[club_name].append(event)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_name = f"{filename_prefix}_{timestamp}"
    temp_root = tempfile.mkdtemp(prefix='club_events_rar_')
    package_root = os.path.join(temp_root, package_name)
    os.makedirs(package_root, exist_ok=True)

    summary = {
        'club_count': 0,
        'event_count': len(events),
        'dynamic_image_count': 0
    }

    try:
        if not grouped_events:
            readme_path = os.path.join(package_root, 'README.txt')
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write("当前时间范围内无活动数据。\n")
        else:
            for club_name in sorted(grouped_events.keys()):
                safe_club_name = sanitize_filename(club_name)
                club_excel_name = f"{safe_club_name}.xlsx"
                club_folder_name = f"{safe_club_name}协会文件夹"

                club_excel_path = os.path.join(package_root, club_excel_name)
                club_asset_folder = os.path.join(package_root, club_folder_name)
                os.makedirs(club_asset_folder, exist_ok=True)

                dynamic_count = create_single_club_event_excel(
                    excel_path=club_excel_path,
                    club_name=club_name,
                    events=grouped_events[club_name],
                    club_asset_folder=club_asset_folder,
                    club_folder_name=club_folder_name,
                    temp_root=temp_root
                )
                summary['dynamic_image_count'] += dynamic_count
                summary['club_count'] += 1

        archive_info = create_export_archive(source_dir=package_root, output_dir=temp_root, package_name=package_name)

        minio_response = upload_local_file_to_minio(
            local_file_path=archive_info['archive_path'],
            object_path=f"statistics/{archive_info['archive_filename']}",
            content_type=archive_info['content_type']
        )

        return jsonify({
            'code': 200,
            'message': archive_info['message'],
            'data': {
                **minio_response,
                **summary,
                'archive_format': archive_info['archive_format'],
                'split_by_club': True,
                'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
                'end_date': end_date.strftime('%Y-%m-%d') if end_date else ''
            }
        })
    finally:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass

def create_all_event_details_single_package(events, filename_prefix, start_date=None, end_date=None):
    """所有活动导出到单个Excel，活动文件夹不按协会分层。"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel支持库未安装")
    if not PILLOW_AVAILABLE:
        raise Exception("图片处理库未安装")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_name = f"{filename_prefix}_single_{timestamp}"
    temp_root = tempfile.mkdtemp(prefix='all_events_single_')
    package_root = os.path.join(temp_root, package_name)
    os.makedirs(package_root, exist_ok=True)

    summary = {
        'club_count': len(set([(event.club.clubID if event.club else -1) for event in events])),
        'event_count': len(events),
        'dynamic_image_count': 0
    }

    try:
        excel_path = os.path.join(package_root, '所有活动.xlsx')
        summary['dynamic_image_count'] = create_all_events_single_excel(
            excel_path=excel_path,
            events=events,
            asset_root=package_root,
            temp_root=temp_root
        )

        archive_info = create_export_archive(source_dir=package_root, output_dir=temp_root, package_name=package_name)
        minio_response = upload_local_file_to_minio(
            local_file_path=archive_info['archive_path'],
            object_path=f"statistics/{archive_info['archive_filename']}",
            content_type=archive_info['content_type']
        )

        return jsonify({
            'code': 200,
            'message': archive_info['message'],
            'data': {
                **minio_response,
                **summary,
                'archive_format': archive_info['archive_format'],
                'split_by_club': False,
                'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
                'end_date': end_date.strftime('%Y-%m-%d') if end_date else ''
            }
        })
    finally:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass

def create_single_club_event_excel(excel_path, club_name, events, club_asset_folder, club_folder_name, temp_root):
    """生成单个协会活动Excel，并把原图放到同级协会文件夹。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(club_name)

    headers = ['活动封面略缩图', '动态图片略缩图', '人员名单', '时间', '地点']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 58
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.freeze_panes = 'A2'

    dynamic_image_count = 0
    temp_thumb_files = []

    sorted_events = sorted(
        events,
        key=lambda item: get_event_reference_datetime(item) or datetime.min,
        reverse=True
    )

    try:
        for row_idx, event in enumerate(sorted_events, 2):
            ws.row_dimensions[row_idx].height = 160
            event_folder_name = f"{sanitize_filename(event.title) or event.eventID}活动"
            event_asset_folder = os.path.join(club_asset_folder, event_folder_name)
            dynamic_asset_folder = os.path.join(event_asset_folder, "动态原图")
            os.makedirs(dynamic_asset_folder, exist_ok=True)

            join_list = EventJoin.query.filter_by(eventID=event.eventID).all()
            names = []
            for join in join_list:
                if hasattr(join, 'isDelete') and join.isDelete:
                    continue
                if join.user and join.user.userName:
                    names.append(f"【{join.user.userName}】")

            member_text = ''.join(names) if names else '无'
            start_text = event.pre_startTime.strftime('%Y-%m-%d %H:%M') if event.pre_startTime else ''
            end_text = event.pre_endTime.strftime('%Y-%m-%d %H:%M') if event.pre_endTime else ''
            time_text = f"{start_text} - {end_text}".strip(' -') or '无'
            location_text = event.location_name or event.location_address or event.location or '无'

            ws.cell(row=row_idx, column=3, value=member_text).alignment = text_alignment
            ws.cell(row=row_idx, column=4, value=time_text).alignment = center_alignment
            ws.cell(row=row_idx, column=5, value=location_text).alignment = text_alignment

            cover_cell = ws.cell(row=row_idx, column=1, value='无封面')
            cover_cell.alignment = center_alignment
            if event.cover and event.cover.fileUrl:
                cover_bytes = download_image_from_minio(event.cover.fileUrl)
                if cover_bytes:
                    original_cover_path = extract_file_path_from_download_url(event.cover.fileUrl)
                    cover_ext = os.path.splitext(original_cover_path)[1] if original_cover_path else '.jpg'
                    cover_file_name = f"cover{cover_ext or '.jpg'}"
                    cover_file_abs = os.path.join(event_asset_folder, cover_file_name)
                    with open(cover_file_abs, 'wb') as f:
                        f.write(cover_bytes)

                    cover_thumb = generate_thumbnail_bytes(cover_bytes, minlength=50)
                    if cover_thumb:
                        fd, thumb_path = tempfile.mkstemp(prefix='cover_thumb_', suffix='.jpg', dir=temp_root)
                        os.close(fd)
                        with open(thumb_path, 'wb') as f:
                            f.write(cover_thumb)
                        temp_thumb_files.append(thumb_path)

                        excel_cover = ExcelImage(thumb_path)
                        excel_cover.anchor = f'A{row_idx}'
                        ws.add_image(excel_cover)

                    relative_cover_link = os.path.relpath(cover_file_abs, os.path.dirname(excel_path)).replace("\\", "/")
                    cover_cell.value = '封面原图'
                    cover_cell.hyperlink = relative_cover_link
                    cover_cell.style = "Hyperlink"

            dynamic_cell = ws.cell(row=row_idx, column=2, value='无动态图片')
            dynamic_cell.alignment = center_alignment
            moment_files = collect_event_moment_files(event)
            if moment_files:
                saved_dynamic_paths = []
                for idx, file_obj in enumerate(moment_files, 1):
                    if not file_obj.fileUrl:
                        continue
                    image_bytes = download_image_from_minio(file_obj.fileUrl)
                    if not image_bytes:
                        continue
                    dynamic_image_count += 1
                    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
                    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
                    dynamic_name = f"moment_{idx}{ext or '.jpg'}"
                    dynamic_abs = os.path.join(dynamic_asset_folder, dynamic_name)
                    with open(dynamic_abs, 'wb') as f:
                        f.write(image_bytes)
                    saved_dynamic_paths.append(dynamic_abs)

                collage_bytes = build_moment_collage_thumbnail(moment_files, minlength=50, max_images=12)
                if collage_bytes:
                    fd, collage_path = tempfile.mkstemp(prefix='moment_thumb_', suffix='.jpg', dir=temp_root)
                    os.close(fd)
                    with open(collage_path, 'wb') as f:
                        f.write(collage_bytes)
                    temp_thumb_files.append(collage_path)

                    excel_collage = ExcelImage(collage_path)
                    excel_collage.anchor = f'B{row_idx}'
                    ws.add_image(excel_collage)

                if saved_dynamic_paths:
                    relative_dynamic_link = os.path.relpath(saved_dynamic_paths[0], os.path.dirname(excel_path)).replace("\\", "/")
                    dynamic_cell.value = '动态原图'
                    dynamic_cell.hyperlink = relative_dynamic_link
                    dynamic_cell.style = "Hyperlink"

        wb.save(excel_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        for tmp_file in temp_thumb_files:
            if tmp_file and os.path.exists(tmp_file):
                try:
                    os.unlink(tmp_file)
                except Exception:
                    pass

    return dynamic_image_count

def create_all_events_single_excel(excel_path, events, asset_root, temp_root):
    """生成所有活动单表Excel，原图文件夹按活动区分。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "全部活动"

    headers = ['协会', '活动封面略缩图', '动态图片略缩图', '人员名单', '时间', '地点']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 58
    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.freeze_panes = 'A2'

    dynamic_image_count = 0
    temp_thumb_files = []
    sorted_events = sorted(
        events,
        key=lambda item: get_event_reference_datetime(item) or datetime.min,
        reverse=True
    )

    try:
        for row_idx, event in enumerate(sorted_events, 2):
            ws.row_dimensions[row_idx].height = 160

            club_name = event.club.clubName if event.club else '未知协会'
            ws.cell(row=row_idx, column=1, value=club_name).alignment = text_alignment

            safe_title = sanitize_filename(event.title) or f"event_{event.eventID}"
            event_folder_name = f"{event.eventID}_{safe_title}活动"
            event_asset_folder = os.path.join(asset_root, event_folder_name)
            dynamic_asset_folder = os.path.join(event_asset_folder, "动态原图")
            os.makedirs(dynamic_asset_folder, exist_ok=True)

            join_list = EventJoin.query.filter_by(eventID=event.eventID).all()
            names = []
            for join in join_list:
                if hasattr(join, 'isDelete') and join.isDelete:
                    continue
                if join.user and join.user.userName:
                    names.append(f"【{join.user.userName}】")

            member_text = ''.join(names) if names else '无'
            start_text = event.pre_startTime.strftime('%Y-%m-%d %H:%M') if event.pre_startTime else ''
            end_text = event.pre_endTime.strftime('%Y-%m-%d %H:%M') if event.pre_endTime else ''
            time_text = f"{start_text} - {end_text}".strip(' -') or '无'
            location_text = event.location_name or event.location_address or event.location or '无'

            ws.cell(row=row_idx, column=4, value=member_text).alignment = text_alignment
            ws.cell(row=row_idx, column=5, value=time_text).alignment = center_alignment
            ws.cell(row=row_idx, column=6, value=location_text).alignment = text_alignment

            cover_cell = ws.cell(row=row_idx, column=2, value='无封面')
            cover_cell.alignment = center_alignment
            if event.cover and event.cover.fileUrl:
                cover_bytes = download_image_from_minio(event.cover.fileUrl)
                if cover_bytes:
                    original_cover_path = extract_file_path_from_download_url(event.cover.fileUrl)
                    cover_ext = os.path.splitext(original_cover_path)[1] if original_cover_path else '.jpg'
                    cover_file_name = f"cover{cover_ext or '.jpg'}"
                    cover_file_abs = os.path.join(event_asset_folder, cover_file_name)
                    with open(cover_file_abs, 'wb') as f:
                        f.write(cover_bytes)

                    cover_thumb = generate_thumbnail_bytes(cover_bytes, minlength=50)
                    if cover_thumb:
                        fd, thumb_path = tempfile.mkstemp(prefix='cover_single_thumb_', suffix='.jpg', dir=temp_root)
                        os.close(fd)
                        with open(thumb_path, 'wb') as f:
                            f.write(cover_thumb)
                        temp_thumb_files.append(thumb_path)

                        excel_cover = ExcelImage(thumb_path)
                        excel_cover.anchor = f'B{row_idx}'
                        ws.add_image(excel_cover)

                    relative_cover_link = os.path.relpath(cover_file_abs, os.path.dirname(excel_path)).replace("\\", "/")
                    cover_cell.value = '封面原图'
                    cover_cell.hyperlink = relative_cover_link
                    cover_cell.style = "Hyperlink"

            dynamic_cell = ws.cell(row=row_idx, column=3, value='无动态图片')
            dynamic_cell.alignment = center_alignment
            moment_files = collect_event_moment_files(event)
            if moment_files:
                saved_dynamic_paths = []
                for idx, file_obj in enumerate(moment_files, 1):
                    if not file_obj.fileUrl:
                        continue
                    image_bytes = download_image_from_minio(file_obj.fileUrl)
                    if not image_bytes:
                        continue
                    dynamic_image_count += 1
                    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
                    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
                    dynamic_name = f"moment_{idx}{ext or '.jpg'}"
                    dynamic_abs = os.path.join(dynamic_asset_folder, dynamic_name)
                    with open(dynamic_abs, 'wb') as f:
                        f.write(image_bytes)
                    saved_dynamic_paths.append(dynamic_abs)

                collage_bytes = build_moment_collage_thumbnail(moment_files, minlength=50, max_images=12)
                if collage_bytes:
                    fd, collage_path = tempfile.mkstemp(prefix='moment_single_thumb_', suffix='.jpg', dir=temp_root)
                    os.close(fd)
                    with open(collage_path, 'wb') as f:
                        f.write(collage_bytes)
                    temp_thumb_files.append(collage_path)

                    excel_collage = ExcelImage(collage_path)
                    excel_collage.anchor = f'C{row_idx}'
                    ws.add_image(excel_collage)

                if saved_dynamic_paths:
                    relative_dynamic_link = os.path.relpath(saved_dynamic_paths[0], os.path.dirname(excel_path)).replace("\\", "/")
                    dynamic_cell.value = '动态原图'
                    dynamic_cell.hyperlink = relative_dynamic_link
                    dynamic_cell.style = "Hyperlink"

        wb.save(excel_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        for tmp_file in temp_thumb_files:
            if tmp_file and os.path.exists(tmp_file):
                try:
                    os.unlink(tmp_file)
                except Exception:
                    pass

    return dynamic_image_count

def create_rar_archive(source_dir, rar_path):
    """使用系统 rar/winrar 打包目录。"""
    rar_commands = []
    rar_exe = shutil.which('rar')
    winrar_exe = shutil.which('winrar')
    if rar_exe:
        rar_commands.append([rar_exe, 'a', '-r', rar_path, '.'])
    if winrar_exe:
        rar_commands.append([winrar_exe, 'a', '-r', rar_path, '.'])

    if not rar_commands:
        raise Exception('服务器未安装 rar/winrar，无法生成RAR压缩包')

    last_error = None
    for cmd in rar_commands:
        try:
            subprocess.run(cmd, cwd=source_dir, check=True, capture_output=True, text=True)
            return
        except Exception as e:
            last_error = e

    raise Exception(f'生成RAR失败: {str(last_error)}')

def create_zip_archive(source_dir, zip_path):
    """使用Python标准库zipfile打包目录。"""
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(source_dir):
            for file_name in files:
                file_path = os.path.join(root, file_name)
                arc_name = os.path.relpath(file_path, os.path.dirname(source_dir))
                zipf.write(file_path, arc_name)

def create_export_archive(source_dir, output_dir, package_name):
    """优先生成RAR，若不可用则自动降级为ZIP。"""
    rar_filename = f"{package_name}.rar"
    rar_path = os.path.join(output_dir, rar_filename)

    try:
        create_rar_archive(source_dir=source_dir, rar_path=rar_path)
        return {
            'archive_path': rar_path,
            'archive_filename': rar_filename,
            'archive_format': 'rar',
            'content_type': 'application/vnd.rar',
            'message': '导出成功'
        }
    except Exception as rar_error:
        current_app.logger.warning(f"RAR打包失败，降级ZIP: {str(rar_error)}")
        zip_filename = f"{package_name}.zip"
        zip_path = os.path.join(output_dir, zip_filename)
        create_zip_archive(source_dir=source_dir, zip_path=zip_path)
        return {
            'archive_path': zip_path,
            'archive_filename': zip_filename,
            'archive_format': 'zip',
            'content_type': 'application/zip',
            'message': '导出成功（服务器未安装RAR工具，已自动提供ZIP压缩包）'
        }

def upload_local_file_to_minio(local_file_path, object_path, content_type):
    """上传本地文件到MinIO并返回下载信息。"""
    minio_client = get_minio_client()
    bucket_name = current_app.config.get('MINIO_BUCKET', 'manage-mate')
    ensure_bucket_exists(minio_client, bucket_name)

    file_size = os.path.getsize(local_file_path)
    with open(local_file_path, 'rb') as f:
        minio_client.put_object(
            bucket_name,
            object_path,
            f,
            length=file_size,
            content_type=content_type
        )

    base_url = (current_app.config.get('BASE_URL') or 'https://www.vhhg.top').rstrip('/')
    download_url = f"{base_url}/api/v1/file/download/tmp/{object_path}"
    return {
        'download_url': download_url,
        'filename': os.path.basename(local_file_path),
        'file_path': object_path,
        'file_size': file_size,
        'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

def get_wecom_access_token():
    """获取企业微信 access_token。"""
    access_token = current_app.config.get('WECOM_TOKEN')
    if access_token:
        return access_token

    corpid = current_app.config.get('WECOM_CORP_ID')
    corpsecret = current_app.config.get('WECOM_SECRET')
    if not corpid or not corpsecret:
        raise Exception('企业微信配置不完整，请配置 WECOM_CORP_ID/WECOM_SECRET')

    token_url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={corpid}&corpsecret={corpsecret}'
    token_response = requests.get(token_url, timeout=15)
    token_data = token_response.json()
    if token_data.get('errcode') != 0:
        raise Exception(f'获取企业微信 access_token 失败: {token_data.get("errmsg", "未知错误")}')

    access_token = token_data.get('access_token')
    if not access_token:
        raise Exception('企业微信 access_token 为空')

    current_app.config['WECOM_TOKEN'] = access_token
    return access_token

def upload_minio_object_to_wecom_media(object_path, file_name):
    """将 MinIO 对象上传到企业微信临时素材，返回 media_id。"""
    access_token = get_wecom_access_token()
    minio_client = get_minio_client()
    bucket_name = current_app.config.get('MINIO_BUCKET', 'manage-mate')

    obj = None
    try:
        obj = minio_client.get_object(bucket_name, object_path)
        file_bytes = obj.read()
    except Exception as e:
        raise Exception(f'读取导出文件失败: {str(e)}')
    finally:
        if obj is not None:
            try:
                obj.close()
                obj.release_conn()
            except Exception:
                pass

    mime_type = mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
    upload_url = f'https://qyapi.weixin.qq.com/cgi-bin/media/upload?access_token={access_token}&type=file'
    upload_resp = requests.post(
        upload_url,
        files={'media': (file_name, file_bytes, mime_type)},
        timeout=60
    )
    upload_data = upload_resp.json()
    if upload_data.get('errcode') != 0 or not upload_data.get('media_id'):
        raise Exception(f'上传企业微信素材失败: {upload_data.get("errmsg", "未知错误")}')

    return upload_data['media_id']

def send_wecom_file_message_to_users(wecom_user_ids, media_id):
    """调用企业微信 message/send，将文件素材发给指定用户列表。"""
    if not wecom_user_ids:
        raise Exception('接收用户为空')
    if not media_id:
        raise Exception('media_id 为空')

    access_token = get_wecom_access_token()
    agentid = current_app.config.get('WECOM_AGENT_ID')
    if not agentid:
        raise Exception('企业微信配置缺少 WECOM_AGENT_ID')

    send_url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
    payload = {
        'touser': '|'.join(wecom_user_ids),
        'msgtype': 'file',
        'agentid': int(agentid),
        'file': {
            'media_id': media_id
        },
        'safe': 0
    }
    resp = requests.post(send_url, json=payload, timeout=30)
    data = resp.json()
    if data.get('errcode') != 0:
        raise Exception(data.get('errmsg', '企业微信 message/send 调用失败'))

def create_all_club_users_archive(clubs, filename_prefix, start_date=None, end_date=None, activity_start_date=None, activity_end_date=None):
    """导出所有协会用户：每协会Excel+协会文件夹，整体打包下载。"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel支持库未安装")
    if not PILLOW_AVAILABLE:
        raise Exception("图片处理库未安装")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_name = f"{filename_prefix}_{timestamp}"
    temp_root = tempfile.mkdtemp(prefix='club_users_archive_')
    package_root = os.path.join(temp_root, package_name)
    os.makedirs(package_root, exist_ok=True)

    summary = {
        'club_count': 0,
        'member_count': 0,
        'moment_image_count': 0
    }

    try:
        if not clubs:
            readme_path = os.path.join(package_root, 'README.txt')
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write("当前没有可导出的协会用户数据。\n")
        else:
            for club in sorted(clubs, key=lambda c: c.clubName or ''):
                safe_club_name = sanitize_filename(club.clubName or f"club_{club.clubID}")
                excel_path = os.path.join(package_root, f"{safe_club_name}.xlsx")
                club_folder_name = f"{safe_club_name}协会文件夹"
                club_asset_folder = os.path.join(package_root, club_folder_name)
                os.makedirs(club_asset_folder, exist_ok=True)

                export_result = create_single_club_user_excel(
                    club=club,
                    excel_path=excel_path,
                    club_asset_folder=club_asset_folder,
                    temp_root=temp_root,
                    start_date=start_date,
                    end_date=end_date,
                    activity_start_date=activity_start_date,
                    activity_end_date=activity_end_date
                )
                summary['club_count'] += 1
                summary['member_count'] += export_result['member_count']
                summary['moment_image_count'] += export_result['moment_image_count']

        archive_info = create_export_archive(source_dir=package_root, output_dir=temp_root, package_name=package_name)
        minio_response = upload_local_file_to_minio(
            local_file_path=archive_info['archive_path'],
            object_path=f"statistics/{archive_info['archive_filename']}",
            content_type=archive_info['content_type']
        )

        return jsonify({
            'code': 200,
            'message': archive_info['message'],
            'data': {
                **minio_response,
                **summary,
                'archive_format': archive_info['archive_format'],
                'split_by_club': True,
                'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
                'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
                'activity_start_date': activity_start_date.strftime('%Y-%m-%d') if activity_start_date else '',
                'activity_end_date': activity_end_date.strftime('%Y-%m-%d') if activity_end_date else ''
            }
        })
    finally:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass

def create_all_users_single_archive(clubs, filename_prefix, start_date=None, end_date=None, activity_start_date=None, activity_end_date=None):
    """导出所有协会用户到单一Excel与单一文件夹。"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel支持库未安装")
    if not PILLOW_AVAILABLE:
        raise Exception("图片处理库未安装")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    package_name = f"{filename_prefix}_single_{timestamp}"
    temp_root = tempfile.mkdtemp(prefix='all_users_single_archive_')
    package_root = os.path.join(temp_root, package_name)
    os.makedirs(package_root, exist_ok=True)

    summary = {
        'club_count': len(clubs),
        'member_count': 0,
        'moment_image_count': 0
    }

    try:
        user_map = {}
        for club in clubs:
            members = ClubMember.query.filter_by(clubID=club.clubID, isDelete=False).all()
            for member in members:
                if not is_member_in_range(member, start_date, end_date):
                    continue
                if not member.user:
                    continue
                uid = member.user.userID
                if uid not in user_map:
                    user_map[uid] = {
                        'user': member.user,
                        'memberships': []
                    }
                user_map[uid]['memberships'].append({
                    'club': club,
                    'member': member
                })

        excel_path = os.path.join(package_root, "所有协会用户.xlsx")
        asset_folder = os.path.join(package_root, "所有协会用户文件夹")
        os.makedirs(asset_folder, exist_ok=True)

        export_result = create_all_users_single_excel(
            excel_path=excel_path,
            asset_folder=asset_folder,
            user_rows=list(user_map.values()),
            temp_root=temp_root,
            activity_start_date=activity_start_date,
            activity_end_date=activity_end_date
        )
        summary['member_count'] = export_result['member_count']
        summary['moment_image_count'] = export_result['moment_image_count']

        archive_info = create_export_archive(source_dir=package_root, output_dir=temp_root, package_name=package_name)
        minio_response = upload_local_file_to_minio(
            local_file_path=archive_info['archive_path'],
            object_path=f"statistics/{archive_info['archive_filename']}",
            content_type=archive_info['content_type']
        )

        return jsonify({
            'code': 200,
            'message': archive_info['message'],
            'data': {
                **minio_response,
                **summary,
                'archive_format': archive_info['archive_format'],
                'split_by_club': False,
                'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
                'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
                'activity_start_date': activity_start_date.strftime('%Y-%m-%d') if activity_start_date else '',
                'activity_end_date': activity_end_date.strftime('%Y-%m-%d') if activity_end_date else ''
            }
        })
    finally:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception:
            pass

def create_all_users_single_excel(excel_path, asset_folder, user_rows, temp_root, activity_start_date=None, activity_end_date=None):
    """生成单表用户导出（新增协会列）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "全部用户"

    headers = ['用户名', '用户头像略缩图', '加入协会时间', '协会', '参加活动', '发布过的动态']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 36
    ws.column_dimensions['E'].width = 56
    ws.column_dimensions['F'].width = 62
    ws.freeze_panes = 'A2'

    sorted_users = sorted(user_rows, key=lambda item: (item['user'].userName or ''))
    temp_thumb_files = []
    moment_image_count = 0

    try:
        for row_idx, row in enumerate(sorted_users, 2):
            user = row['user']
            memberships = row['memberships']
            club_names = [m['club'].clubName for m in memberships if m['club'] and m['club'].clubName]
            club_ids = [m['club'].clubID for m in memberships if m['club']]

            ws.row_dimensions[row_idx].height = 120
            safe_user_name = sanitize_filename(user.userName or f"user_{user.userID}")
            folder_prefix = sanitize_filename(club_names[0]) if len(club_names) == 1 else f"{sanitize_filename(club_names[0])}等{len(club_names)}协会"
            user_folder_name = f"{folder_prefix}-{safe_user_name}文件夹" if club_names else f"未知协会-{safe_user_name}文件夹"
            user_asset_folder = os.path.join(asset_folder, user_folder_name)
            moment_asset_folder = os.path.join(user_asset_folder, '动态原图')
            os.makedirs(moment_asset_folder, exist_ok=True)

            ws.cell(row=row_idx, column=1, value=user.userName or '').alignment = text_alignment

            join_dates = [m['member'].joinDate for m in memberships if m['member'] and m['member'].joinDate]
            earliest_join = min(join_dates).strftime('%Y-%m-%d %H:%M:%S') if join_dates else ''
            ws.cell(row=row_idx, column=3, value=earliest_join).alignment = center_alignment

            clubs_text = ''.join([f"【{name}】" for name in sorted(set(club_names))]) if club_names else '无'
            ws.cell(row=row_idx, column=4, value=clubs_text).alignment = text_alignment

            event_names = get_user_joined_event_names(
                user.userID,
                club_ids,
                include_club_prefix=True,
                start_date=activity_start_date,
                end_date=activity_end_date
            )
            events_text = ''.join([f"【{name}】" for name in event_names]) if event_names else '无'
            ws.cell(row=row_idx, column=5, value=events_text).alignment = text_alignment

            avatar_cell = ws.cell(row=row_idx, column=2, value='无头像')
            avatar_cell.alignment = center_alignment
            if user.avatar and user.avatar.fileUrl:
                avatar_bytes = download_image_from_minio(user.avatar.fileUrl)
                if avatar_bytes:
                    avatar_origin_path = extract_file_path_from_download_url(user.avatar.fileUrl)
                    avatar_ext = os.path.splitext(avatar_origin_path)[1] if avatar_origin_path else '.jpg'
                    avatar_name = f"avatar{avatar_ext or '.jpg'}"
                    avatar_abs_path = os.path.join(user_asset_folder, avatar_name)
                    with open(avatar_abs_path, 'wb') as f:
                        f.write(avatar_bytes)

                    avatar_thumb = generate_thumbnail_bytes(avatar_bytes, minlength=30)
                    if avatar_thumb:
                        fd, thumb_path = tempfile.mkstemp(prefix='avatar_single_thumb_', suffix='.jpg', dir=temp_root)
                        os.close(fd)
                        with open(thumb_path, 'wb') as f:
                            f.write(avatar_thumb)
                        temp_thumb_files.append(thumb_path)
                        avatar_img = ExcelImage(thumb_path)
                        avatar_img.anchor = f'B{row_idx}'
                        ws.add_image(avatar_img)

                    avatar_cell.value = '头像原图'
                    avatar_cell.hyperlink = os.path.relpath(avatar_abs_path, os.path.dirname(excel_path)).replace("\\", "/")
                    avatar_cell.style = "Hyperlink"

            moments_cell = ws.cell(row=row_idx, column=6, value='无动态')
            moments_cell.alignment = center_alignment
            moment_files = collect_user_moment_files(
                user.userID,
                club_ids,
                start_date=activity_start_date,
                end_date=activity_end_date
            )
            if moment_files:
                saved_paths = []
                for idx, file_obj in enumerate(moment_files, 1):
                    if not file_obj.fileUrl:
                        continue
                    image_bytes = download_image_from_minio(file_obj.fileUrl)
                    if not image_bytes:
                        continue
                    moment_image_count += 1
                    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
                    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
                    image_name = f"moment_{idx}{ext or '.jpg'}"
                    image_abs = os.path.join(moment_asset_folder, image_name)
                    with open(image_abs, 'wb') as f:
                        f.write(image_bytes)
                    saved_paths.append(image_abs)

                collage_bytes = build_moment_collage_thumbnail(moment_files, minlength=50, max_images=20)
                if collage_bytes:
                    fd, collage_path = tempfile.mkstemp(prefix='all_user_moment_thumb_', suffix='.jpg', dir=temp_root)
                    os.close(fd)
                    with open(collage_path, 'wb') as f:
                        f.write(collage_bytes)
                    temp_thumb_files.append(collage_path)
                    collage_img = ExcelImage(collage_path)
                    collage_img.anchor = f'F{row_idx}'
                    ws.add_image(collage_img)

                if saved_paths:
                    moments_cell.value = '动态原图'
                    moments_cell.hyperlink = os.path.relpath(saved_paths[0], os.path.dirname(excel_path)).replace("\\", "/")
                    moments_cell.style = "Hyperlink"

        wb.save(excel_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        for temp_file in temp_thumb_files:
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except Exception:
                    pass

    return {
        'member_count': len(sorted_users),
        'moment_image_count': moment_image_count
    }

def create_single_club_user_excel(club, excel_path, club_asset_folder, temp_root, start_date=None, end_date=None, activity_start_date=None, activity_end_date=None):
    """生成单个协会用户导出Excel与原图目录。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(club.clubName or f"club_{club.clubID}")

    headers = ['用户名', '用户头像略缩图', '加入协会时间', '参加活动', '发布过的动态']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 48
    ws.column_dimensions['E'].width = 62
    ws.freeze_panes = 'A2'

    members = ClubMember.query.filter_by(clubID=club.clubID, isDelete=False).all()
    members = [member for member in members if is_member_in_range(member, start_date, end_date)]
    temp_thumb_files = []
    moment_image_count = 0

    try:
        for row_idx, member in enumerate(members, 2):
            user = member.user
            if not user:
                continue

            ws.row_dimensions[row_idx].height = 120
            safe_user_name = sanitize_filename(user.userName or f"user_{user.userID}")
            user_folder_name = f"{sanitize_filename(club.clubName)}-{safe_user_name}文件夹"
            user_asset_folder = os.path.join(club_asset_folder, user_folder_name)
            moment_asset_folder = os.path.join(user_asset_folder, '动态原图')
            os.makedirs(moment_asset_folder, exist_ok=True)

            ws.cell(row=row_idx, column=1, value=user.userName or '').alignment = text_alignment
            ws.cell(
                row=row_idx,
                column=3,
                value=member.joinDate.strftime('%Y-%m-%d %H:%M:%S') if member.joinDate else ''
            ).alignment = center_alignment

            event_names = get_user_joined_event_names(
                user.userID,
                [club.clubID],
                include_club_prefix=False,
                start_date=activity_start_date,
                end_date=activity_end_date
            )
            ws.cell(row=row_idx, column=4, value=''.join([f"【{name}】" for name in event_names]) if event_names else '无').alignment = text_alignment

            avatar_cell = ws.cell(row=row_idx, column=2, value='无头像')
            avatar_cell.alignment = center_alignment
            if user.avatar and user.avatar.fileUrl:
                avatar_bytes = download_image_from_minio(user.avatar.fileUrl)
                if avatar_bytes:
                    avatar_origin_path = extract_file_path_from_download_url(user.avatar.fileUrl)
                    avatar_ext = os.path.splitext(avatar_origin_path)[1] if avatar_origin_path else '.jpg'
                    avatar_file_name = f"avatar{avatar_ext or '.jpg'}"
                    avatar_abs_path = os.path.join(user_asset_folder, avatar_file_name)
                    with open(avatar_abs_path, 'wb') as f:
                        f.write(avatar_bytes)

                    avatar_thumb = generate_thumbnail_bytes(avatar_bytes, minlength=30)
                    if avatar_thumb:
                        fd, thumb_path = tempfile.mkstemp(prefix='avatar_thumb_', suffix='.jpg', dir=temp_root)
                        os.close(fd)
                        with open(thumb_path, 'wb') as f:
                            f.write(avatar_thumb)
                        temp_thumb_files.append(thumb_path)
                        avatar_img = ExcelImage(thumb_path)
                        avatar_img.anchor = f'B{row_idx}'
                        ws.add_image(avatar_img)

                    avatar_cell.value = '头像原图'
                    avatar_cell.hyperlink = os.path.relpath(avatar_abs_path, os.path.dirname(excel_path)).replace("\\", "/")
                    avatar_cell.style = "Hyperlink"

            moments_cell = ws.cell(row=row_idx, column=5, value='无动态')
            moments_cell.alignment = center_alignment
            moment_files = collect_user_moment_files(
                user.userID,
                [club.clubID],
                start_date=activity_start_date,
                end_date=activity_end_date
            )
            if moment_files:
                saved_paths = []
                for idx, file_obj in enumerate(moment_files, 1):
                    if not file_obj.fileUrl:
                        continue
                    image_bytes = download_image_from_minio(file_obj.fileUrl)
                    if not image_bytes:
                        continue
                    moment_image_count += 1
                    origin_path = extract_file_path_from_download_url(file_obj.fileUrl)
                    ext = os.path.splitext(origin_path)[1] if origin_path else '.jpg'
                    file_name = f"moment_{idx}{ext or '.jpg'}"
                    file_abs = os.path.join(moment_asset_folder, file_name)
                    with open(file_abs, 'wb') as f:
                        f.write(image_bytes)
                    saved_paths.append(file_abs)

                collage_bytes = build_moment_collage_thumbnail(moment_files, minlength=50, max_images=16)
                if collage_bytes:
                    fd, collage_path = tempfile.mkstemp(prefix='user_moment_thumb_', suffix='.jpg', dir=temp_root)
                    os.close(fd)
                    with open(collage_path, 'wb') as f:
                        f.write(collage_bytes)
                    temp_thumb_files.append(collage_path)
                    collage_img = ExcelImage(collage_path)
                    collage_img.anchor = f'E{row_idx}'
                    ws.add_image(collage_img)

                if saved_paths:
                    moments_cell.value = '动态原图'
                    moments_cell.hyperlink = os.path.relpath(saved_paths[0], os.path.dirname(excel_path)).replace("\\", "/")
                    moments_cell.style = "Hyperlink"

        wb.save(excel_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        for temp_file in temp_thumb_files:
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except Exception:
                    pass

    return {
        'member_count': len(members),
        'moment_image_count': moment_image_count
    }

def get_user_joined_event_names(user_id, club_ids, include_club_prefix=False, start_date=None, end_date=None):
    """获取用户在指定协会集合中的活动名称（可选包含协会前缀）。"""
    if not club_ids:
        return []

    normalized_club_ids = [int(cid) for cid in club_ids if cid is not None]
    if not normalized_club_ids:
        return []

    joins = EventJoin.query.join(Event, Event.eventID == EventJoin.eventID).join(
        Club, Club.clubID == Event.clubID
    ).filter(
        EventJoin.userID == user_id,
        Event.clubID.in_(normalized_club_ids)
    ).all()

    names = []
    for join in joins:
        if hasattr(join, 'isDelete') and join.isDelete:
            continue
        if not join.event or not join.event.title:
            continue
        if start_date and end_date and not is_event_in_range(join.event, start_date, end_date):
            continue
        if include_club_prefix and join.event.club and join.event.club.clubName:
            name = f"{join.event.club.clubName}-{join.event.title}"
        else:
            name = join.event.title
        if name not in names:
            names.append(name)
    return names


def query_user_moments_in_clubs(user_id, club_ids, start_date=None, end_date=None):
    """查询用户在多个协会下的动态列表。"""
    if not club_ids:
        return []

    normalized_club_ids = [int(cid) for cid in club_ids if cid is not None]
    if not normalized_club_ids:
        return []

    club_event_ids = [
        event.eventID for event in Event.query.filter(Event.clubID.in_(normalized_club_ids)).all()
    ]
    moment_conditions = [Moment.ref_club_ID.in_(normalized_club_ids)]
    if club_event_ids:
        moment_conditions.append(Moment.ref_event_ID.in_(club_event_ids))

    moments = Moment.query.filter(
        Moment.creatorID == user_id,
        db.or_(*moment_conditions)
    ).order_by(Moment.createDate.asc()).all()

    if start_date and end_date:
        moments = [
            moment for moment in moments
            if moment.createDate and start_date <= moment.createDate <= end_date
        ]
    return moments


def collect_user_moment_files(user_id, club_ids, start_date=None, end_date=None):
    """获取用户在多个协会发布过的动态图片文件。"""
    from app.models.file import File

    moments = query_user_moments_in_clubs(
        user_id=user_id,
        club_ids=club_ids,
        start_date=start_date,
        end_date=end_date,
    )

    image_ids = []
    for moment in moments:
        if moment.imageIDs and isinstance(moment.imageIDs, list):
            for image_id in moment.imageIDs:
                if image_id not in image_ids:
                    image_ids.append(image_id)

    if not image_ids:
        return []

    files = File.query.filter(File.fileID.in_(image_ids)).all()
    file_map = {f.fileID: f for f in files}
    return [file_map[file_id] for file_id in image_ids if file_id in file_map]


def count_user_moments_in_clubs(user_id, club_ids, start_date=None, end_date=None):
    """统计用户在多个协会中的动态数量（按动态时间可选过滤）。"""
    moments = query_user_moments_in_clubs(
        user_id=user_id,
        club_ids=club_ids,
        start_date=start_date,
        end_date=end_date,
    )
    return len(moments)

def download_image_from_minio(image_url):
    """从MinIO下载图片数据并转换为Excel支持的格式（fileUrl 须为当前 BASE_URL 或相对下载路径）。"""
    try:
        image_url = (image_url or '').strip()
        file_path = extract_file_path_from_download_url(image_url)
        if not file_path:
            current_app.logger.error(f"无效的图片URL格式: {image_url}")
            return None
        
        # 获取MinIO客户端
        minio_client = get_minio_client()
        bucket_name = current_app.config.get('MINIO_BUCKET', 'manage-mate')
        
        # 下载图片数据
        response = minio_client.get_object(bucket_name, file_path)
        image_data = response.read()
        response.close()
        response.release_conn()
        
        # 检查图片格式并转换为Excel支持的格式
        if PILLOW_AVAILABLE:
            img_stream = None
            output_stream = None
            try:
                # 使用Pillow打开图片
                img_stream = io.BytesIO(image_data)
                pil_image = PILImage.open(img_stream)
                
                # 检查图片格式
                image_format = pil_image.format
                current_app.logger.info(f"图片格式: {image_format}")
                
                # 如果是不支持的格式（如webp），转换为JPEG
                if image_format and image_format.upper() in ['WEBP', 'BMP', 'TIFF']:
                    current_app.logger.info(f"转换图片格式从 {image_format} 到 JPEG")
                    # 如果图片有透明通道，需要转换为RGB
                    if pil_image.mode in ('RGBA', 'LA', 'P'):
                        # 创建白色背景
                        background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                        if pil_image.mode == 'P':
                            pil_image = pil_image.convert('RGBA')
                        background.paste(pil_image, mask=pil_image.split()[-1] if pil_image.mode == 'RGBA' else None)
                        pil_image = background
                    elif pil_image.mode != 'RGB':
                        pil_image = pil_image.convert('RGB')
                    
                    # 转换为JPEG格式
                    output_stream = io.BytesIO()
                    pil_image.save(output_stream, format='JPEG', quality=85)
                    image_data = output_stream.getvalue()
                    current_app.logger.info(f"成功转换图片格式为JPEG")
                
            except Exception as conversion_error:
                current_app.logger.error(f"图片格式转换失败: {str(conversion_error)}")
                # 如果转换失败，返回原始数据，让Excel处理
                pass
            finally:
                # 确保所有流都被正确关闭
                if img_stream:
                    try:
                        img_stream.close()
                    except:
                        pass
                if output_stream:
                    try:
                        output_stream.close()
                    except:
                        pass
        
        current_app.logger.info(f"成功下载图片: {file_path}")
        return image_data
        
    except Exception as e:
        current_app.logger.error(f"从MinIO下载图片失败 {image_url}: {str(e)}")
        return None

@bp.route('/export/test_image_support', methods=['GET'])
@jwt_required()
def test_image_support():
    """测试图片导出功能的支持情况"""
    # 权限检查
    has_permission, message = check_permission(statistics.test_image_support.permission_judge)
    if not has_permission:
        return jsonify({'code': 4003, 'message': message}), 200
    
    user_id = get_jwt_identity()
    current_user = User.query.filter_by(userID=user_id).first()
    
    if not current_user:
        return jsonify({'code': 4004, 'message': '用户不存在'}), 200
    
    support_info = {
        'excel_available': EXCEL_AVAILABLE,
        'pillow_available': PILLOW_AVAILABLE,
        'image_export_supported': EXCEL_AVAILABLE and PILLOW_AVAILABLE
    }
    
    if EXCEL_AVAILABLE and PILLOW_AVAILABLE:
        message = '图片导出功能完全支持'
        code = 200
    elif EXCEL_AVAILABLE:
        message = '仅支持基础Excel导出，缺少图片处理库Pillow'
        code = 200
    else:
        message = '不支持Excel导出功能，缺少openpyxl库'
        code = 500
    
    return jsonify({
        'code': code,
        'message': message,
        'data': support_info
    })
